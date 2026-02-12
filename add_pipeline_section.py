"""
Add BESS Pipeline & Saturation Analysis Section to HTML Report & Excel
======================================================================
Corrected: Pipeline is 1.7 GW / 4.0 GWh (not 4 GW).
Adds pipeline analysis, saturation timeline, and updated market intelligence.
"""

import os
import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

DATA_DIR = "/Users/mayk/LithuaniaBESS/data"
OUT_DIR = "/Users/mayk/LithuaniaBESS"

# Styling
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
API_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="1F4E79")

# ============================================================
# Pipeline Data
# ============================================================

# Corrected: 1.7 GW / 4.0 GWh pipeline (avg ~2.4h duration)
PIPELINE_TOTAL_MW = 1700
PIPELINE_TOTAL_MWH = 4000
PIPELINE_AVG_DURATION = PIPELINE_TOTAL_MWH / PIPELINE_TOTAL_MW  # ~2.35h

# Key projects from research + LinkedIn
KNOWN_PROJECTS = [
    {'developer': 'UAB Vėjo Galia', 'mw': 53.6, 'mwh': 107.3, 'location': 'Kaišiadorys',
     'status': 'Operational', 'year': 2025, 'source': 'Litgrid (first commercial BESS on transmission grid)'},
    {'developer': 'European Energy', 'mw': 25, 'mwh': 65, 'location': 'Anykščiai',
     'status': 'Construction complete', 'year': 2026, 'source': 'LinkedIn Jan 2026'},
    {'developer': 'Litgrid (TSO-owned)', 'mw': 200, 'mwh': 200, 'location': 'Various',
     'status': 'Operational', 'year': 2024, 'source': 'First large-capacity facility, 2020 commissioned'},
    {'developer': 'Ignitis Group', 'mw': 130, 'mwh': 260, 'location': 'TBD',
     'status': 'Development', 'year': 2026, 'source': 'EUR 130M investment announced'},
    {'developer': 'E Energija Group', 'mw': 100, 'mwh': 200, 'location': 'TBD',
     'status': 'Under construction', 'year': 2026, 'source': 'LinkedIn (construction underway)'},
    {'developer': 'Fluence / Litgrid', 'mw': 50, 'mwh': 100, 'location': 'TBD',
     'status': 'Integration', 'year': 2025, 'source': 'System integrator collaboration'},
]

# Actual market sizing from ENTSO-E data
print("Loading market data for pipeline analysis...")

afrr = pd.read_csv(f"{DATA_DIR}/afrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
afrr.index = pd.to_datetime(afrr.index, utc=True)
for c in afrr.columns:
    afrr[c] = pd.to_numeric(afrr[c], errors='coerce')

mfrr = pd.read_csv(f"{DATA_DIR}/mfrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
mfrr.index = pd.to_datetime(mfrr.index, utc=True)
for c in mfrr.columns:
    mfrr[c] = pd.to_numeric(mfrr[c], errors='coerce')

da = pd.read_csv(f"{DATA_DIR}/da_prices_LT.csv", index_col=0, parse_dates=True)
da.columns = ['price']
da.index = pd.to_datetime(da.index, utc=True)
da['price'] = pd.to_numeric(da['price'], errors='coerce')

# 2025 procurement volumes
mask25 = afrr.index.year == 2025
afrr_up_mean = afrr.loc[mask25, 'Up Quantity'].mean()
afrr_down_mean = afrr.loc[mask25, 'Down Quantity'].mean()
afrr_up_max = afrr.loc[mask25, 'Up Quantity'].max()

mask25m = mfrr.index.year == 2025
mfrr_up_mean = mfrr.loc[mask25m, 'Up Quantity'].mean()
mfrr_down_mean = mfrr.loc[mask25m, 'Down Quantity'].mean()

total_balancing = afrr_up_mean + afrr_down_mean + mfrr_up_mean + mfrr_down_mean
fcr_est = 40  # MW estimated

# DA metrics
mask25_da = da.index.year == 2025
da_daily = da.loc[mask25_da, 'price'].groupby(da.loc[mask25_da].index.date)
da_spread_mean = (da_daily.max() - da_daily.min()).mean()
neg_hours = (da.loc[mask25_da, 'price'] < 0).sum()
high_hours = (da.loc[mask25_da, 'price'] > 200).sum()

print(f"  aFRR Up: {afrr_up_mean:.0f} MW, Down: {afrr_down_mean:.0f} MW")
print(f"  mFRR Up: {mfrr_up_mean:.0f} MW, Down: {mfrr_down_mean:.0f} MW")
print(f"  Total balancing: {total_balancing:.0f} MW")
print(f"  DA spread: {da_spread_mean:.0f} EUR/MWh, Neg hours: {neg_hours}, >200 EUR hours: {high_hours}")

# Build-out timeline scenarios
SCENARIOS = {
    'High': {
        2025: 454, 2026: 800, 2027: 1200, 2028: 1500, 2029: 1700, 2030: 1700
    },
    'Base': {
        2025: 454, 2026: 650, 2027: 950, 2028: 1200, 2029: 1400, 2030: 1500
    },
    'Low': {
        2025: 454, 2026: 550, 2027: 700, 2028: 850, 2029: 1000, 2030: 1100
    }
}

# Saturation metrics per scenario
PEAK_LOAD = 2100  # MW
ANNUAL_CONSUMPTION = 13.4  # TWh

print("\n=== Pipeline Saturation Analysis ===")
for scenario, timeline in SCENARIOS.items():
    print(f"\n  {scenario} scenario:")
    for year, mw in timeline.items():
        sat_afrr = mw / afrr_up_mean * 100
        sat_total = mw / (total_balancing + fcr_est) * 100
        sat_load = mw / PEAK_LOAD * 100
        print(f"    {year}: {mw:>5} MW → aFRR:{sat_afrr:5.0f}%, Balancing:{sat_total:5.0f}%, Load:{sat_load:5.0f}%")


# ============================================================
# UPDATE EXCEL
# ============================================================
print("\n\nUpdating Excel workbook...")
XLSX_PATH = f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx"

def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def write_data_cell(ws, row, col, val, fmt=None, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    cell.fill = fill or API_FILL
    if fmt:
        cell.number_format = fmt

def auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)

wb = load_workbook(XLSX_PATH)

if "Pipeline & Saturation" in wb.sheetnames:
    del wb["Pipeline & Saturation"]

ws = wb.create_sheet("Pipeline & Saturation", 3)

row = 1
ws.cell(row=row, column=1, value="Lithuania BESS Pipeline & Market Saturation Analysis").font = TITLE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
ws.cell(row=row, column=1,
        value="Pipeline: 1.7 GW / 4.0 GWh | 50+ applications | EUR 840M+ total value | 14.7% state subsidy").font = NOTE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 2

# ---- Pipeline overview ----
ws.cell(row=row, column=1, value="Pipeline Overview").font = SUBTITLE_FONT
row += 1
overview_data = [
    ('Total Pipeline Power', f'{PIPELINE_TOTAL_MW:,} MW'),
    ('Total Pipeline Energy', f'{PIPELINE_TOTAL_MWH:,} MWh'),
    ('Average Duration', f'{PIPELINE_AVG_DURATION:.1f} hours'),
    ('Number of Applications', '50+'),
    ('Total Project Value', 'EUR 840M+'),
    ('State Subsidy (average)', '14.7% of costs'),
    ('Max per Facility', '300 MWh'),
    ('Currently Installed', '454 MW / 462 MWh'),
    ('', ''),
    ('Key Entities', ''),
    ('Ministry of Energy', 'Policy & procurement framework'),
    ('EPMA', 'Environmental Project Management Agency'),
    ('Litgrid (TSO)', '200 MW operational + grid integration'),
    ('Ignitis Group', 'EUR 130M BESS investment'),
    ('E Energija Group', 'Under construction'),
    ('European Energy', '25 MW / 65 MWh completed Jan 2026 (Anykščiai)'),
    ('UAB Vėjo Galia', '53.6 MW / 107.3 MWh operational (Kaišiadorys)'),
    ('Fluence', 'System integrator with Litgrid'),
    ('UAB Karjerų Linija', 'Investment Jul 2025'),
]
for label, val in overview_data:
    ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=bool(label), size=10)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=val).font = DATA_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')
    row += 1

row += 1

# ---- Market Size vs Pipeline ----
ws.cell(row=row, column=1, value="Market Size vs Pipeline").font = SUBTITLE_FONT
row += 1
write_header_row(ws, row, ['Market Segment', 'Actual (MW)', 'Pipeline as Multiple',
                            'Saturation Level', 'Implication'])
row += 1
market_vs_pipe = [
    ('aFRR Up Capacity', f'{afrr_up_mean:.0f}', f'{PIPELINE_TOTAL_MW/afrr_up_mean:.1f}x',
     'Oversupplied', 'Revenue compression 2027+'),
    ('aFRR Down Capacity', f'{afrr_down_mean:.0f}', f'{PIPELINE_TOTAL_MW/afrr_down_mean:.1f}x',
     'Oversupplied', 'Revenue compression 2027+'),
    ('mFRR Up + Down', f'{mfrr_up_mean + mfrr_down_mean:.0f}', f'{PIPELINE_TOTAL_MW/(mfrr_up_mean+mfrr_down_mean):.1f}x',
     'Heavily oversupplied', 'Sporadic revenue only'),
    ('FCR (estimated)', f'{fcr_est}', f'{PIPELINE_TOTAL_MW/fcr_est:.0f}x',
     'Heavily oversupplied', 'Small market segment'),
    ('Total Balancing', f'{total_balancing + fcr_est:.0f}',
     f'{PIPELINE_TOTAL_MW/(total_balancing + fcr_est):.1f}x',
     'Oversupplied', 'Balancing cannot absorb full pipeline'),
    ('Peak Load', f'{PEAK_LOAD}', f'{PIPELINE_TOTAL_MW/PEAK_LOAD:.1f}x',
     'High penetration', 'DA arbitrage structurally viable'),
]
for label, actual, mult, sat, impl in market_vs_pipe:
    write_data_cell(ws, row, 1, label)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    write_data_cell(ws, row, 2, actual)
    write_data_cell(ws, row, 3, mult)
    sat_fill = WARN_FILL if 'Over' in sat else HIGHLIGHT_FILL
    write_data_cell(ws, row, 4, sat, fill=sat_fill)
    write_data_cell(ws, row, 5, impl)
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='left')
    row += 1

row += 1

# ---- Build-out Scenarios ----
ws.cell(row=row, column=1, value="BESS Build-Out Scenarios (Installed MW)").font = SUBTITLE_FONT
row += 1
scenario_start = row
write_header_row(ws, row, ['Year', 'High', 'Base', 'Low',
                            'High % aFRR Up', 'Base % aFRR Up', 'Low % aFRR Up',
                            'aFRR Up Req. (MW)'])
row += 1
for year in range(2025, 2031):
    write_data_cell(ws, row, 1, year)
    for i, (name, sc) in enumerate(SCENARIOS.items()):
        mw = sc[year]
        write_data_cell(ws, row, 2 + i, mw, fmt='#,##0')
        pct = mw / afrr_up_mean * 100
        fill = WARN_FILL if pct > 200 else (HIGHLIGHT_FILL if pct > 100 else API_FILL)
        write_data_cell(ws, row, 5 + i, f'{pct:.0f}%', fill=fill)
    write_data_cell(ws, row, 8, round(afrr_up_mean))
    row += 1

# Chart for scenarios
chart = LineChart()
chart.title = "BESS Build-Out Scenarios vs aFRR Requirement"
chart.y_axis.title = "Installed MW"
chart.x_axis.title = "Year"
chart.style = 10
chart.width = 22
chart.height = 13

cats = Reference(ws, min_col=1, min_row=scenario_start + 1, max_row=scenario_start + 6)
for col_idx, name in [(2, 'High'), (3, 'Base'), (4, 'Low')]:
    data_ref = Reference(ws, min_col=col_idx, min_row=scenario_start, max_row=scenario_start + 6)
    chart.add_data(data_ref, titles_from_data=True)

# Add aFRR line
afrr_ref = Reference(ws, min_col=8, min_row=scenario_start, max_row=scenario_start + 6)
chart.add_data(afrr_ref, titles_from_data=True)

chart.set_categories(cats)
ws.add_chart(chart, f"A{row + 1}")
row += 18

# ---- Key Projects ----
ws.cell(row=row, column=1, value="Known BESS Projects").font = SUBTITLE_FONT
row += 1
write_header_row(ws, row, ['Developer', 'MW', 'MWh', 'Duration (h)',
                            'Location', 'Status', 'Year', 'Source'])
row += 1
for proj in KNOWN_PROJECTS:
    dur = proj['mwh'] / proj['mw'] if proj['mw'] > 0 else 0
    write_data_cell(ws, row, 1, proj['developer'])
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    write_data_cell(ws, row, 2, proj['mw'], fmt='#,##0')
    write_data_cell(ws, row, 3, proj['mwh'], fmt='#,##0')
    write_data_cell(ws, row, 4, round(dur, 1), fmt='0.0')
    write_data_cell(ws, row, 5, proj['location'])
    write_data_cell(ws, row, 6, proj['status'])
    write_data_cell(ws, row, 7, proj['year'])
    write_data_cell(ws, row, 8, proj['source'])
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='left', wrap_text=True)
    row += 1

row += 1

# ---- Conversion probability ----
ws.cell(row=row, column=1, value="Pipeline Conversion Analysis").font = SUBTITLE_FONT
row += 1
conv_notes = [
    "Pipeline: 1.7 GW / 4.0 GWh across 50+ applications (EUR 840M+, avg 14.7% state subsidy).",
    "Connection rights ≠ built projects. In mature EU markets, 20-40% conversion from application to FID.",
    "At 30% conversion: ~510 MW new capacity → total ~960 MW by 2028.",
    "At 60% conversion: ~1,020 MW new capacity → total ~1,470 MW by 2029.",
    "At 90% conversion: ~1,530 MW new → total ~1,980 MW by 2030 (if grid & supply chain allow).",
    "Limiting factors: grid connection capacity, transformer availability, LFP cell supply chain, financing.",
    "State subsidy (14.7% avg) de-risks projects but initial funding was doubled (demand >> budget).",
    "Max 300 MWh per facility limits concentration risk but increases grid connection complexity.",
    "European Energy (25 MW/65 MWh) Anykščiai project completed Jan 2026 — ~2.6h duration, portfolio approach.",
    "First-mover advantage: projects online 2025-2026 earn extraordinary returns before saturation.",
]
for note in conv_notes:
    ws.cell(row=row, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

auto_width(ws)
ws.column_dimensions['A'].width = 30
ws.column_dimensions['H'].width = 45

wb.save(XLSX_PATH)
print(f"  Excel updated: added 'Pipeline & Saturation' sheet")


# ============================================================
# UPDATE HTML REPORT
# ============================================================
print("\nUpdating HTML report...")

HTML_PATH = f"{OUT_DIR}/Lithuania_BESS_Market_Report.html"
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

# Scenario data for Plotly
years_list = list(range(2025, 2031))
scenario_traces = {}
for name, sc in SCENARIOS.items():
    scenario_traces[name] = [sc[y] for y in years_list]
afrr_line = [round(afrr_up_mean)] * len(years_list)
total_bal_line = [round(total_balancing + fcr_est)] * len(years_list)

# Saturation waterfall data
waterfall_labels = ['aFRR Up', 'aFRR Down', 'mFRR', 'FCR (est.)', 'DA Arbitrage\n(structural)']
waterfall_values = [
    round(afrr_up_mean),
    round(afrr_down_mean),
    round(mfrr_up_mean + mfrr_down_mean),
    fcr_est,
    PEAK_LOAD  # DA opportunity proportional to load
]

# Projects table HTML
projects_html = ""
for proj in KNOWN_PROJECTS:
    dur = proj['mwh'] / proj['mw'] if proj['mw'] > 0 else 0
    status_color = {
        'Operational': '#27ae60',
        'Construction complete': '#2980b9',
        'Under construction': '#f39c12',
        'Development': '#e67e22',
        'Integration': '#8e44ad'
    }.get(proj['status'], '#666')
    projects_html += f"""<tr>
        <td style="text-align:left">{proj['developer']}</td>
        <td>{proj['mw']}</td><td>{proj['mwh']}</td><td>{dur:.1f}</td>
        <td>{proj['location']}</td>
        <td><span style="background:{status_color};color:white;padding:2px 8px;border-radius:10px;font-size:0.85em">{proj['status']}</span></td>
        <td>{proj['year']}</td>
    </tr>"""

# Conversion probability table
conv_data = [
    (20, 340, 454 + 340, '2029-2030'),
    (30, 510, 454 + 510, '2028-2029'),
    (40, 680, 454 + 680, '2028'),
    (60, 1020, 454 + 1020, '2027-2028'),
    (80, 1360, 454 + 1360, '2027'),
    (100, 1700, 454 + 1700, '2028-2030'),
]
conv_html = ""
for pct, new_mw, total, timeline in conv_data:
    sat_afrr = total / afrr_up_mean * 100
    sat_class = 'color:#c0392b;font-weight:bold' if sat_afrr > 200 else ('color:#e67e22' if sat_afrr > 100 else 'color:#27ae60')
    conv_html += f"""<tr>
        <td>{pct}%</td>
        <td>{new_mw:,} MW</td>
        <td>{total:,} MW</td>
        <td style="{sat_class}">{sat_afrr:.0f}%</td>
        <td>{timeline}</td>
    </tr>"""

new_section = f"""
<!-- ========== PIPELINE & SATURATION ========== -->
<div style="margin-top:50px;">
    <h2 style="color:#1F4E79; border-bottom:3px solid #1F4E79; padding-bottom:10px;">
        4. BESS Pipeline & Market Saturation Analysis
    </h2>

    <!-- Pipeline headline -->
    <div style="background:linear-gradient(135deg, #2c3e50 0%, #3498db 100%); color:white; padding:30px; border-radius:12px; margin:20px 0;">
        <div style="display:grid; grid-template-columns:repeat(4, 1fr); gap:20px; text-align:center;">
            <div>
                <div style="font-size:2.2em; font-weight:bold;">1.7 GW</div>
                <div style="color:#bdc3c7; font-size:0.9em;">Pipeline Power</div>
            </div>
            <div>
                <div style="font-size:2.2em; font-weight:bold;">4.0 GWh</div>
                <div style="color:#bdc3c7; font-size:0.9em;">Pipeline Energy</div>
            </div>
            <div>
                <div style="font-size:2.2em; font-weight:bold;">454 MW</div>
                <div style="color:#bdc3c7; font-size:0.9em;">Installed Today</div>
            </div>
            <div>
                <div style="font-size:2.2em; font-weight:bold;">50+</div>
                <div style="color:#bdc3c7; font-size:0.9em;">Applications</div>
            </div>
        </div>
        <div style="display:grid; grid-template-columns:repeat(4, 1fr); gap:20px; text-align:center; margin-top:15px; padding-top:15px; border-top:1px solid rgba(255,255,255,0.2);">
            <div>
                <div style="font-size:1.4em; font-weight:bold;">EUR 840M+</div>
                <div style="color:#bdc3c7; font-size:0.85em;">Total Investment</div>
            </div>
            <div>
                <div style="font-size:1.4em; font-weight:bold;">14.7%</div>
                <div style="color:#bdc3c7; font-size:0.85em;">Avg State Subsidy</div>
            </div>
            <div>
                <div style="font-size:1.4em; font-weight:bold;">~2.4h</div>
                <div style="color:#bdc3c7; font-size:0.85em;">Avg Duration</div>
            </div>
            <div>
                <div style="font-size:1.4em; font-weight:bold;">300 MWh</div>
                <div style="color:#bdc3c7; font-size:0.85em;">Max per Facility</div>
            </div>
        </div>
    </div>

    <!-- Market Size vs Pipeline -->
    <h3 style="color:#2E75B6; margin-top:35px;">Pipeline vs Addressable Market</h3>
    <p style="color:#555; font-size:0.93em;">
        How does 1.7 GW of pipeline compare to actual market demand? Based on 2025 ENTSO-E procurement data:
    </p>
    <div id="pipelineWaterfall" style="width:100%; height:450px;"></div>

    <div style="overflow-x:auto; margin-top:15px;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px; text-align:left">Market Segment</th>
                <th style="padding:10px">2025 Actual (MW)</th>
                <th style="padding:10px">Pipeline Multiple</th>
                <th style="padding:10px">Assessment</th>
            </tr>
        </thead>
        <tbody>
            <tr><td style="text-align:left">aFRR Up Capacity</td><td>{afrr_up_mean:.0f}</td><td><strong>{PIPELINE_TOTAL_MW/afrr_up_mean:.1f}x</strong></td><td style="color:#e67e22">Oversupplied if fully built</td></tr>
            <tr><td style="text-align:left">aFRR Down Capacity</td><td>{afrr_down_mean:.0f}</td><td>{PIPELINE_TOTAL_MW/afrr_down_mean:.1f}x</td><td style="color:#e67e22">Oversupplied if fully built</td></tr>
            <tr><td style="text-align:left">mFRR Up + Down</td><td>{mfrr_up_mean+mfrr_down_mean:.0f}</td><td>{PIPELINE_TOTAL_MW/(mfrr_up_mean+mfrr_down_mean):.1f}x</td><td style="color:#c0392b">Heavily oversupplied</td></tr>
            <tr><td style="text-align:left">FCR (estimated)</td><td>~{fcr_est}</td><td>{PIPELINE_TOTAL_MW/fcr_est:.0f}x</td><td style="color:#c0392b">Small market</td></tr>
            <tr style="font-weight:bold; background:#eee;"><td style="text-align:left">Total Balancing</td><td>{total_balancing+fcr_est:.0f}</td><td>{PIPELINE_TOTAL_MW/(total_balancing+fcr_est):.1f}x</td><td style="color:#e67e22">Cannot absorb full pipeline</td></tr>
            <tr><td style="text-align:left">Peak System Load</td><td>{PEAK_LOAD:,}</td><td>{PIPELINE_TOTAL_MW/PEAK_LOAD:.1f}x</td><td style="color:#27ae60">DA arbitrage structurally viable</td></tr>
        </tbody>
    </table>
    </div>

    <!-- Build-out Scenarios -->
    <h3 style="color:#2E75B6; margin-top:35px;">Build-Out Scenarios</h3>
    <p style="color:#555; font-size:0.93em;">
        Not all 1.7 GW will be built. Connection rights &#8800; operational projects. In mature EU markets,
        20-40% of pipeline typically reaches FID. Three scenarios model different conversion rates:
    </p>
    <div id="scenarioChart" style="width:100%; height:500px;"></div>

    <!-- Conversion probability table -->
    <h3 style="color:#2E75B6; margin-top:35px;">Pipeline Conversion Probability</h3>
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px">Conversion Rate</th>
                <th style="padding:10px">New Capacity</th>
                <th style="padding:10px">Total Installed</th>
                <th style="padding:10px">% of aFRR Up</th>
                <th style="padding:10px">Likely Timeline</th>
            </tr>
        </thead>
        <tbody>
            {conv_html}
        </tbody>
    </table>
    </div>

    <!-- Known Projects -->
    <h3 style="color:#2E75B6; margin-top:35px;">Known BESS Projects</h3>
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px; text-align:left">Developer</th>
                <th style="padding:10px">MW</th>
                <th style="padding:10px">MWh</th>
                <th style="padding:10px">Duration</th>
                <th style="padding:10px">Location</th>
                <th style="padding:10px">Status</th>
                <th style="padding:10px">Year</th>
            </tr>
        </thead>
        <tbody>
            {projects_html}
        </tbody>
    </table>
    </div>

    <!-- Investment Timing Analysis -->
    <h3 style="color:#2E75B6; margin-top:35px;">Investment Timing: The Window of Opportunity</h3>
    <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:16px; margin:20px 0;">
        <div style="background:#e8f5e9; border:2px solid #27ae60; border-radius:8px; padding:20px;">
            <h4 style="color:#27ae60; margin:0 0 10px 0;">2025-2026: Golden Window</h4>
            <ul style="font-size:0.9em; color:#333; padding-left:18px; margin:0;">
                <li>454 MW installed vs 876 MW demand</li>
                <li>Post-BRELL scarcity pricing</li>
                <li>aFRR Up: EUR 29.56/MW mean</li>
                <li>Sub-1-year payback possible</li>
                <li>Limited competition</li>
            </ul>
        </div>
        <div style="background:#fff3e0; border:2px solid #f39c12; border-radius:8px; padding:20px;">
            <h4 style="color:#f39c12; margin:0 0 10px 0;">2027-2028: Compression Zone</h4>
            <ul style="font-size:0.9em; color:#333; padding-left:18px; margin:0;">
                <li>~950-1,200 MW installed (Base)</li>
                <li>aFRR market approaching saturation</li>
                <li>Revenue 45-65% of 2025 levels</li>
                <li>2-3 year payback</li>
                <li>Multi-market strategy essential</li>
            </ul>
        </div>
        <div style="background:#fce4ec; border:2px solid #c0392b; border-radius:8px; padding:20px;">
            <h4 style="color:#c0392b; margin:0 0 10px 0;">2029-2030: Saturated Market</h4>
            <ul style="font-size:0.9em; color:#333; padding-left:18px; margin:0;">
                <li>1,400-1,700 MW installed</li>
                <li>Balancing fully saturated</li>
                <li>DA arbitrage still viable (structural)</li>
                <li>Revenue 28-35% of 2025</li>
                <li>Only low-CAPEX projects viable</li>
            </ul>
        </div>
    </div>

    <!-- Structural tailwinds -->
    <h3 style="color:#2E75B6; margin-top:35px;">Structural Tailwinds (Offsetting Saturation)</h3>
    <div style="background:#f8f9fa; padding:20px; border-radius:8px; font-size:0.92em; color:#444;">
        <div style="display:grid; grid-template-columns:1fr 1fr; gap:20px;">
            <div>
                <strong style="color:#1F4E79;">Growing Renewables</strong>
                <p style="margin:4px 0;">Lithuania targeting 5.1 GW solar and doubled wind by 2030 (NECP).
                1.7 GW added in 2025 alone. More renewables = more price volatility = more BESS opportunity.
                DA daily spread already EUR {da_spread_mean:.0f} mean in 2025.</p>
            </div>
            <div>
                <strong style="color:#1F4E79;">Interconnection Risks</strong>
                <p style="margin:4px 0;">Estlink 2 cable failure caused 65-71% Baltic price surge in Feb 2025.
                Baltics remain an "energy island" — interconnection outages create scarcity events
                where BESS captures extreme prices (EUR 9,976/MWh spikes observed).</p>
            </div>
            <div>
                <strong style="color:#1F4E79;">Baltic Market Integration</strong>
                <p style="margin:4px 0;">Joint Baltic balancing capacity procurement launched Feb 2025.
                Lithuania BESS can serve Latvian/Estonian reserves, expanding addressable market beyond
                Lithuania-only procurement. Combined Baltic aFRR need likely 600-800 MW.</p>
            </div>
            <div>
                <strong style="color:#1F4E79;">Reserve Capacity Fees</strong>
                <p style="margin:4px 0;">Estonia introduced EUR 3.70/MWh reserve fee for all consumers post-BRELL.
                Similar mechanisms across Baltics create a permanent revenue pool for flexibility providers,
                independent of spot market dynamics.</p>
            </div>
        </div>
    </div>
</div>
"""

# Plotly scripts
new_scripts = f"""
<script>
(function() {{
    // Waterfall: market segments vs pipeline
    var waterfall = {{
        type: 'bar',
        x: ['aFRR Up', 'aFRR Down', 'mFRR', 'FCR (est.)', 'Total Balancing'],
        y: [{round(afrr_up_mean)}, {round(afrr_down_mean)}, {round(mfrr_up_mean+mfrr_down_mean)}, {fcr_est}, {round(total_balancing+fcr_est)}],
        name: 'Market Demand (MW)',
        marker: {{color: ['#2E75B6', '#2E75B6', '#2E75B6', '#2E75B6', '#1F4E79']}}
    }};
    var pipeline_line = {{
        type: 'scatter',
        mode: 'lines',
        x: ['aFRR Up', 'aFRR Down', 'mFRR', 'FCR (est.)', 'Total Balancing'],
        y: [{PIPELINE_TOTAL_MW}, {PIPELINE_TOTAL_MW}, {PIPELINE_TOTAL_MW}, {PIPELINE_TOTAL_MW}, {PIPELINE_TOTAL_MW}],
        name: 'Pipeline: 1,700 MW',
        line: {{color: '#c0392b', width: 3, dash: 'dash'}}
    }};
    var installed_line = {{
        type: 'scatter',
        mode: 'lines',
        x: ['aFRR Up', 'aFRR Down', 'mFRR', 'FCR (est.)', 'Total Balancing'],
        y: [454, 454, 454, 454, 454],
        name: 'Installed: 454 MW',
        line: {{color: '#27ae60', width: 3, dash: 'dot'}}
    }};
    Plotly.newPlot('pipelineWaterfall', [waterfall, pipeline_line, installed_line], {{
        title: 'BESS Pipeline (1.7 GW) vs Actual Market Demand',
        yaxis: {{title: 'MW'}},
        legend: {{orientation: 'h', y: -0.15}},
        margin: {{t: 50, b: 80}},
        plot_bgcolor: '#fafafa',
        annotations: [{{
            x: 'Total Balancing', y: {PIPELINE_TOTAL_MW},
            text: '1.7 GW pipeline = {PIPELINE_TOTAL_MW/(total_balancing+fcr_est):.1f}x total balancing',
            showarrow: true, arrowhead: 2, ax: -80, ay: -40,
            font: {{size: 11, color: '#c0392b'}}
        }}]
    }}, {{responsive: true, displayModeBar: false}});

    // Scenario chart
    var years = {json.dumps([str(y) for y in years_list])};
    var traces = [
        {{name: 'High Scenario', x: years, y: {json.dumps(scenario_traces['High'])},
          mode: 'lines+markers', line: {{color: '#c0392b', width: 2}}, marker: {{size: 8}}}},
        {{name: 'Base Scenario', x: years, y: {json.dumps(scenario_traces['Base'])},
          mode: 'lines+markers', line: {{color: '#f39c12', width: 3}}, marker: {{size: 8}}}},
        {{name: 'Low Scenario', x: years, y: {json.dumps(scenario_traces['Low'])},
          mode: 'lines+markers', line: {{color: '#27ae60', width: 2}}, marker: {{size: 8}}}},
        {{name: 'aFRR Up Requirement', x: years, y: {json.dumps(afrr_line)},
          mode: 'lines', line: {{color: '#2E75B6', width: 2, dash: 'dash'}}}},
        {{name: 'Total Balancing Need', x: years, y: {json.dumps(total_bal_line)},
          mode: 'lines', line: {{color: '#1F4E79', width: 2, dash: 'dot'}}}}
    ];
    Plotly.newPlot('scenarioChart', traces, {{
        title: 'BESS Build-Out Scenarios vs Balancing Market Size',
        yaxis: {{title: 'MW', range: [0, 1900]}},
        legend: {{orientation: 'h', y: -0.15}},
        margin: {{t: 50, b: 80}},
        plot_bgcolor: '#fafafa',
        shapes: [{{
            type: 'rect', x0: '2025', x1: '2026', y0: 0, y1: 1900,
            fillcolor: 'rgba(39,174,96,0.07)', line: {{width: 0}},
            layer: 'below'
        }}, {{
            type: 'rect', x0: '2027', x1: '2028', y0: 0, y1: 1900,
            fillcolor: 'rgba(243,156,18,0.07)', line: {{width: 0}},
            layer: 'below'
        }}, {{
            type: 'rect', x0: '2029', x1: '2030', y0: 0, y1: 1900,
            fillcolor: 'rgba(192,57,43,0.07)', line: {{width: 0}},
            layer: 'below'
        }}],
        annotations: [
            {{x: '2025', y: 100, text: 'Golden<br>Window', showarrow: false, font: {{size: 10, color: '#27ae60'}}}},
            {{x: '2027', y: 100, text: 'Compression<br>Zone', showarrow: false, font: {{size: 10, color: '#f39c12'}}}},
            {{x: '2029', y: 100, text: 'Saturated<br>Market', showarrow: false, font: {{size: 10, color: '#c0392b'}}}}
        ]
    }}, {{responsive: true, displayModeBar: false}});
}})();
</script>
"""

# Inject section — after revenue analysis, before extended balancing
insert_marker = '<!-- ========== EXTENDED BALANCING DATA ========== -->'
if insert_marker in html:
    html = html.replace(insert_marker, new_section + '\n' + insert_marker)
else:
    # Try before Load & Generation
    alt_marker = '<!-- ========== LOAD & GENERATION ========== -->'
    if alt_marker in html:
        html = html.replace(alt_marker, new_section + '\n' + alt_marker)
    else:
        html = html.replace('</div><!-- container -->', new_section + '\n</div><!-- container -->')

html = html.replace('</body>', new_scripts + '\n</body>')

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"  HTML report updated with Pipeline & Saturation section")


# ============================================================
# UPDATE README with corrected figures
# ============================================================
print("\nUpdating README.md...")

readme_path = f"{OUT_DIR}/README.md"
with open(readme_path, 'r') as f:
    readme = f.read()

# Fix the pipeline figure in README
readme = readme.replace('~4 GW approved/planned', '1.7 GW / 4.0 GWh pipeline (50+ applications)')
readme = readme.replace(
    '**BESS Pipeline** | ~4 GW approved/planned',
    '**BESS Pipeline** | 1.7 GW / 4.0 GWh (50+ applications, EUR 840M+)'
)

# Add pipeline script to table
if 'add_pipeline_section.py' not in readme:
    readme = readme.replace(
        '| `add_revenue_section.py` | Add BESS revenue analysis by duration and market |',
        '| `add_revenue_section.py` | Add BESS revenue analysis by duration and market |\n'
        '| `add_pipeline_section.py` | Add pipeline & saturation analysis with build-out scenarios |'
    )

# Add step 7
if 'add_pipeline_section.py' not in readme:
    readme = readme.replace(
        '# 6. Open report',
        '# 6. Add pipeline & saturation analysis\npython add_pipeline_section.py\n\n# 7. Open report'
    )

with open(readme_path, 'w') as f:
    f.write(readme)

print(f"  README.md updated with corrected pipeline figures")

print("\n" + "=" * 60)
print("DONE — Pipeline & Saturation Analysis Added")
print("=" * 60)
print(f"  Pipeline: {PIPELINE_TOTAL_MW:,} MW / {PIPELINE_TOTAL_MWH:,} MWh (avg {PIPELINE_AVG_DURATION:.1f}h)")
print(f"  vs aFRR Up: {afrr_up_mean:.0f} MW → {PIPELINE_TOTAL_MW/afrr_up_mean:.1f}x oversupply")
print(f"  vs Total Balancing: {total_balancing+fcr_est:.0f} MW → {PIPELINE_TOTAL_MW/(total_balancing+fcr_est):.1f}x oversupply")
print(f"  vs Peak Load: {PEAK_LOAD:,} MW → {PIPELINE_TOTAL_MW/PEAK_LOAD:.1f}x")
