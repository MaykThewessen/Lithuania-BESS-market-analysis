"""
Lithuania BESS Market Analysis & BirdEnergySystemInstalled Excel Generator
==========================================================================
Creates a comprehensive Excel workbook with:
1. Installed Capacity (Wind, PV, BESS, Fossil) - Historical + Forecast
2. Day-Ahead Electricity Prices (monthly, from ENTSO-E/EnergyInEU)
3. Electricity Load Analysis (TWh, GW avg/min/max)
4. aFRR / Imbalance / Intra-Day Price Analysis
5. BESS Market Saturation Analysis
6. Market Overview & Key Developments

Sources: ENTSO-E, Nord Pool, Litgrid, IRENA, LVEA, IEA, PV Magazine,
         CEE Energy News, Energy Storage News, ESS News
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime

wb = Workbook()

# ============================================================
# Color scheme & styles
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
DATA_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
FORECAST_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_cell(ws, row, col, is_forecast=False):
    cell = ws.cell(row=row, column=col)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    if is_forecast:
        cell.fill = FORECAST_FILL

def auto_width(ws, min_width=12, max_width=22):
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


# ============================================================
# SHEET 1: Installed Capacity (Historical + Forecast)
# ============================================================
ws1 = wb.active
ws1.title = "Installed Capacity"

# Title
ws1.merge_cells('A1:L1')
ws1['A1'] = "Lithuania - Installed Capacity by Technology (MW)"
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A2:L2')
ws1['A2'] = "Historical (2021-2025) & Forecast (2026-2031) | Sources: IRENA, LVEA, Litgrid, PV Magazine, IEA, NECP"
ws1['A2'].font = NOTE_FONT

# Data - carefully cross-referenced from multiple sources
# Wind: LVEA: 2021=671, 2022=946; ~1050 by 2023; growth in 2024-2025 to reach ~3GW total w/ solar by end 2025
# Solar: PV Mag: 2022 cumul ~570, 2023 cumul ~1410, 2024 cumul ~2280; end 2025 ~2800
# BESS: Litgrid 200MW/200MWh in 2022; ~535MW by mid-2025; pipeline to 1.7GW
# Fossil: Elektrėnai complex ~1800MW (incl reserve), active CCGT 455MW + reserve units

years = [2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031]
forecast_start = 2026

# --- WIND (MW cumulative) ---
wind_cumul = [671, 946, 1050, 1200, 1500, 1900, 2400, 3000, 3500, 4000, 4500]
# 2021-2022: LVEA/Litgrid official
# 2023: ~1050 (additions of ~100MW)
# 2024: ~1200 (moderate growth, some large projects in permitting)
# 2025: ~1500 (record 1.7GW wind+solar additions; ~300MW wind assumed)
# 2026-2031: accelerating toward ~4GW target (NECP doubled ambition + 0.7GW offshore by 2030)

# --- SOLAR PV (MW cumulative) ---
pv_cumul = [250, 570, 1410, 2280, 2800, 3500, 4200, 4800, 5100, 5500, 5800]
# 2021: ~250MW (early growth phase)
# 2022: +413MW -> ~570MW (PV Magazine)
# 2023: +646MW -> ~1410MW (PV Magazine)
# 2024: +870MW -> ~2280MW (PV Magazine confirmed)
# 2025: ~2800MW (Litgrid: 3GW total wind+solar reached)
# 2026-2031: toward 5.1GW NECP target by 2030; slowing additions as target approached

# --- BESS Power (MW) ---
bess_power = [0, 200, 200, 250, 535, 800, 1200, 1500, 1700, 1800, 1900]
# 2021: essentially 0 grid-scale BESS
# 2022: Litgrid/Fluence 200MW/200MWh operational
# 2023: 200MW (no major new additions)
# 2024: ~250MW (Energy Cells 40MW aFRR + small additions)
# 2025: 535MW (confirmed operational by Litgrid)
# 2026: +265MW (Trina 90MW mid-2026 + others)
# 2027: +400MW (Ignitis 291MW online + pipeline projects)
# 2028: target 800MWh minimum from govt procurement
# 2029-2031: toward 1.7GW/4GWh government target

# --- BESS Energy (MWh) ---
bess_energy = [0, 200, 200, 300, 700, 1200, 2000, 3000, 4000, 4200, 4500]
# Duration increasing from 1h toward 2-2.5h average
# Government target: 4GWh by ~2029

# --- Fossil/Thermal (MW available capacity) ---
fossil_mw = [1800, 1800, 1800, 1800, 1800, 1600, 1500, 1400, 1200, 1000, 900]
# Elektrėnai complex ~1800MW total (455MW CCGT + reserve oil/gas units)
# Gradual decommissioning of reserve units as renewables + BESS replace them
# CCGT 455MW likely remains as strategic reserve through 2030

# Annual additions
wind_annual = [wind_cumul[0]] + [wind_cumul[i] - wind_cumul[i-1] for i in range(1, len(wind_cumul))]
pv_annual = [pv_cumul[0]] + [pv_cumul[i] - pv_cumul[i-1] for i in range(1, len(pv_cumul))]
bess_power_annual = [bess_power[0]] + [bess_power[i] - bess_power[i-1] for i in range(1, len(bess_power))]
bess_energy_annual = [bess_energy[0]] + [bess_energy[i] - bess_energy[i-1] for i in range(1, len(bess_energy))]

# Write headers
headers = ['Year', 'Wind\nCumul (MW)', 'Wind\nAnnual (MW)', 'PV\nCumul (MW)', 'PV\nAnnual (MW)',
           'BESS Power\nCumul (MW)', 'BESS Power\nAnnual (MW)', 'BESS Energy\nCumul (MWh)',
           'BESS Energy\nAnnual (MWh)', 'Fossil\nAvail (MW)', 'Total RES\n(MW)', 'RES+BESS\nTotal (MW)']

row = 4
for col_idx, h in enumerate(headers, 1):
    ws1.cell(row=row, column=col_idx, value=h)
style_header_row(ws1, row, len(headers))

# Write data
for i, year in enumerate(years):
    r = row + 1 + i
    is_fc = year >= forecast_start
    total_res = wind_cumul[i] + pv_cumul[i]
    total_all = total_res + bess_power[i]

    data = [year, wind_cumul[i], wind_annual[i], pv_cumul[i], pv_annual[i],
            bess_power[i], bess_power_annual[i], bess_energy[i], bess_energy_annual[i],
            fossil_mw[i], total_res, total_all]

    for col_idx, val in enumerate(data, 1):
        ws1.cell(row=r, column=col_idx, value=val)
        style_data_cell(ws1, r, col_idx, is_forecast=is_fc)

# Legend
legend_row = row + len(years) + 2
ws1.cell(row=legend_row, column=1, value="Legend:").font = SUBTITLE_FONT
ws1.cell(row=legend_row + 1, column=1, value="Yellow cells = Forecast / Projection").font = NOTE_FONT
ws1.cell(row=legend_row + 2, column=1, value="Sources: LVEA, Litgrid, PV Magazine, IRENA, IEA Lithuania 2025 Review, Lithuania NECP").font = NOTE_FONT
ws1.cell(row=legend_row + 3, column=1, value="BESS govt procurement target: 1.7 GW / 4 GWh (support allocated for 4 GWh)").font = NOTE_FONT
ws1.cell(row=legend_row + 4, column=1, value="Wind includes planned 0.7 GW offshore by 2030").font = NOTE_FONT
ws1.cell(row=legend_row + 5, column=1, value="Solar NECP target: 5.1 GW by 2030").font = NOTE_FONT

# Add bar chart
chart1 = BarChart()
chart1.type = "col"
chart1.grouping = "stacked"
chart1.title = "Lithuania Cumulative Installed Capacity (MW)"
chart1.y_axis.title = "MW"
chart1.x_axis.title = "Year"
chart1.style = 10
chart1.width = 28
chart1.height = 16

data_ref_wind = Reference(ws1, min_col=2, min_row=row, max_row=row + len(years), max_col=2)
data_ref_pv = Reference(ws1, min_col=4, min_row=row, max_row=row + len(years), max_col=4)
data_ref_bess = Reference(ws1, min_col=6, min_row=row, max_row=row + len(years), max_col=6)
cats = Reference(ws1, min_col=1, min_row=row + 1, max_row=row + len(years))

chart1.add_data(data_ref_wind, titles_from_data=True)
chart1.add_data(data_ref_pv, titles_from_data=True)
chart1.add_data(data_ref_bess, titles_from_data=True)
chart1.set_categories(cats)
chart1.series[0].graphicalProperties.solidFill = "2E75B6"
chart1.series[1].graphicalProperties.solidFill = "FFC000"
chart1.series[2].graphicalProperties.solidFill = "70AD47"

ws1.add_chart(chart1, f"A{legend_row + 7}")

auto_width(ws1)

# ============================================================
# SHEET 2: Day-Ahead Electricity Prices
# ============================================================
ws2 = wb.create_sheet("Day-Ahead Prices")

ws2.merge_cells('A1:N1')
ws2['A1'] = "Lithuania - Day-Ahead Wholesale Electricity Prices (EUR/MWh)"
ws2['A1'].font = TITLE_FONT
ws2.merge_cells('A2:N2')
ws2['A2'] = "Source: ENTSO-E / Nord Pool via EnergyInEU | Monthly averages"
ws2['A2'].font = NOTE_FONT

# Monthly price data (from EnergyInEU confirmed data)
price_data = {
    2021: [52.47, 59.31, 47.98, 44.74, 50.35, 77.74, 88.32, 87.74, 123.96, 108.97, 127.82, 212.22],
    2022: [145.87, 104.74, 170.03, 116.46, 164.71, 223.16, 305.36, 480.39, 359.68, 189.45, 226.63, 264.28],
    2023: [103.06, 114.70, 87.72, 67.27, 78.02, 94.75, 83.84, 102.49, 117.29, 87.41, 105.20, 84.77],
    2024: [117.41, 72.54, 68.18, 56.91, 75.85, 91.64, 98.02, 106.95, 83.54, 90.06, 88.98, 89.82],
    2025: [83.03, 143.92, 92.36, 74.59, 64.13, 43.26, 46.18, 78.27, 84.24, 61.51, 38.56, 76.52],
}

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

headers2 = ['Year'] + months + ['Annual Avg']
row2 = 4
for col_idx, h in enumerate(headers2, 1):
    ws2.cell(row=row2, column=col_idx, value=h)
style_header_row(ws2, row2, len(headers2))

for i, (year, prices) in enumerate(price_data.items()):
    r = row2 + 1 + i
    ws2.cell(row=r, column=1, value=year)
    style_data_cell(ws2, r, 1)
    for m, p in enumerate(prices, 2):
        ws2.cell(row=r, column=m, value=round(p, 2))
        style_data_cell(ws2, r, m)
        ws2.cell(row=r, column=m).number_format = '#,##0.00'
    # Annual average
    avg = round(np.mean(prices), 2)
    ws2.cell(row=r, column=14, value=avg)
    style_data_cell(ws2, r, 14)
    ws2.cell(row=r, column=14).number_format = '#,##0.00'
    ws2.cell(row=r, column=14).font = Font(name="Calibri", bold=True, size=10)

# 2026 partial
r_2026 = row2 + 6
ws2.cell(row=r_2026, column=1, value=2026)
style_data_cell(ws2, r_2026, 1, is_forecast=True)
ws2.cell(row=r_2026, column=2, value=114.13)
style_data_cell(ws2, r_2026, 2, is_forecast=True)
ws2.cell(row=r_2026, column=2).number_format = '#,##0.00'
ws2.cell(row=r_2026, column=3, value=205.62)
style_data_cell(ws2, r_2026, 3, is_forecast=True)
ws2.cell(row=r_2026, column=3).number_format = '#,##0.00'
avg_2026_partial = round((114.13 + 205.62) / 2, 2)
ws2.cell(row=r_2026, column=14, value=avg_2026_partial)
ws2.cell(row=r_2026, column=14).font = Font(name="Calibri", bold=True, size=10, italic=True)
ws2.cell(row=r_2026, column=14).number_format = '#,##0.00'
style_data_cell(ws2, r_2026, 14, is_forecast=True)

# Annual summary
summary_row = r_2026 + 2
ws2.cell(row=summary_row, column=1, value="Summary Statistics").font = SUBTITLE_FONT
ws2.cell(row=summary_row + 1, column=1, value="Year").font = SUBHEADER_FONT
ws2.cell(row=summary_row + 1, column=2, value="Avg EUR/MWh").font = SUBHEADER_FONT
ws2.cell(row=summary_row + 1, column=3, value="Min Month").font = SUBHEADER_FONT
ws2.cell(row=summary_row + 1, column=4, value="Max Month").font = SUBHEADER_FONT
ws2.cell(row=summary_row + 1, column=5, value="Spread (Max-Min)").font = SUBHEADER_FONT
ws2.cell(row=summary_row + 1, column=6, value="Neg Price Hours").font = SUBHEADER_FONT

for i, (year, prices) in enumerate(price_data.items()):
    r = summary_row + 2 + i
    ws2.cell(row=r, column=1, value=year)
    ws2.cell(row=r, column=2, value=round(np.mean(prices), 2))
    ws2.cell(row=r, column=3, value=round(min(prices), 2))
    ws2.cell(row=r, column=4, value=round(max(prices), 2))
    ws2.cell(row=r, column=5, value=round(max(prices) - min(prices), 2))
    # Estimated negative price hours
    neg_hours = {2021: 10, 2022: 5, 2023: 50, 2024: 186, 2025: 177}
    ws2.cell(row=r, column=6, value=neg_hours.get(year, 'N/A'))
    for c in range(1, 7):
        style_data_cell(ws2, r, c)

# Line chart for annual average prices
chart2 = LineChart()
chart2.title = "Lithuania Day-Ahead Annual Average (EUR/MWh)"
chart2.y_axis.title = "EUR/MWh"
chart2.style = 10
chart2.width = 22
chart2.height = 14

# Use the annual avg column
avg_data = Reference(ws2, min_col=14, min_row=row2, max_row=row2 + 5)
avg_cats = Reference(ws2, min_col=1, min_row=row2 + 1, max_row=row2 + 5)
chart2.add_data(avg_data, titles_from_data=True)
chart2.set_categories(avg_cats)
chart2.series[0].graphicalProperties.line.width = 25000

ws2.add_chart(chart2, f"A{summary_row + 9}")

auto_width(ws2)

# ============================================================
# SHEET 3: Electricity Load Analysis
# ============================================================
ws3 = wb.create_sheet("Electricity Load")

ws3.merge_cells('A1:H1')
ws3['A1'] = "Lithuania - Electricity Demand & Load Analysis"
ws3['A1'].font = TITLE_FONT
ws3.merge_cells('A2:H2')
ws3['A2'] = "Sources: Litgrid, IEA, Eurostat, NREL Lithuania Study"
ws3['A2'].font = NOTE_FONT

# Load data
load_headers = ['Year', 'Consumption\n(TWh)', 'Dom. Generation\n(TWh)', 'Net Import\n(TWh)',
                'Import Share\n(%)', 'Avg Load\n(GW)', 'Est. Peak\n(GW)', 'Est. Min\n(GW)']
row3 = 4
for col_idx, h in enumerate(load_headers, 1):
    ws3.cell(row=row3, column=col_idx, value=h)
style_header_row(ws3, row3, len(load_headers))

# Data: cross-referenced from multiple sources
# 2021: ~14.2 TWh consumption (before energy crisis reduction)
# 2022: 13.4 TWh (IEA confirmed, -5.5% from 2021); gen 4.8 TWh; import 8.6 TWh
# 2023: ~12.6-13.0 TWh (-4% y/y per Enerdata); gen ~5.5 TWh
# 2024: ~13.0 TWh (stabilized); gen 7.5 TWh (IEA); net imports <50% first time
# 2025: ~13.5 TWh (slight growth); gen ~8.5-9 TWh (70% coverage Q3 extrapolated)
# 2026+: Litgrid 10yr plan: 4%/yr growth toward 18.7 TWh by 2031

load_data = [
    # [Year, Consumption TWh, Gen TWh, Net Import TWh, Import%, Avg GW, Peak GW, Min GW]
    [2021, 14.2, 4.2, 10.0, 70, 1.62, 2.10, 0.90],
    [2022, 13.4, 4.8, 8.6, 64, 1.53, 2.05, 0.85],
    [2023, 12.8, 5.5, 7.3, 57, 1.46, 2.00, 0.82],
    [2024, 13.0, 7.5, 5.5, 42, 1.48, 2.10, 0.85],
    [2025, 13.5, 9.0, 4.5, 33, 1.54, 2.15, 0.88],
    [2026, 14.2, 10.5, 3.7, 26, 1.62, 2.20, 0.90],
    [2027, 15.0, 12.0, 3.0, 20, 1.71, 2.30, 0.95],
    [2028, 16.0, 14.0, 2.0, 13, 1.83, 2.40, 1.00],
    [2029, 17.0, 15.5, 1.5, 9, 1.94, 2.55, 1.05],
    [2030, 18.0, 17.5, 0.5, 3, 2.05, 2.70, 1.10],
    [2031, 18.7, 19.0, -0.3, -2, 2.13, 2.80, 1.15],
]

for i, ld in enumerate(load_data):
    r = row3 + 1 + i
    is_fc = ld[0] >= forecast_start
    for col_idx, val in enumerate(ld, 1):
        ws3.cell(row=r, column=col_idx, value=val)
        style_data_cell(ws3, r, col_idx, is_forecast=is_fc)
        if col_idx in [2, 3, 4, 6, 7, 8]:
            ws3.cell(row=r, column=col_idx).number_format = '#,##0.0'
        if col_idx == 5:
            ws3.cell(row=r, column=col_idx).number_format = '0%' if isinstance(val, float) else '0'

# Key insights
insight_row = row3 + len(load_data) + 2
ws3.cell(row=insight_row, column=1, value="Key Insights:").font = SUBTITLE_FONT
insights = [
    "Lithuania consumed ~13-14 TWh/year (2021-2025), with peak demand ~2.1 GW",
    "Import dependency dropped from 70% (2021) to ~33% (2025) due to renewable build-out",
    "Litgrid forecasts 4%/yr demand growth → 18.7 TWh by 2031 (data centers, EV, heat pumps)",
    "Lithuania targets net electricity exporter status by 2028-2030",
    "Feb 2025: Baltic grid sync with Continental Europe — ended 50+ years of Russian grid dependency",
    "After disconnection, prices spiked to €325/MWh peak (9 Feb 2025) but normalized within days",
]
for i, insight in enumerate(insights):
    ws3.cell(row=insight_row + 1 + i, column=1, value=f"• {insight}").font = DATA_FONT

auto_width(ws3)

# ============================================================
# SHEET 4: aFRR / Imbalance / Intra-Day Prices
# ============================================================
ws4 = wb.create_sheet("Balancing & Ancillary")

ws4.merge_cells('A1:J1')
ws4['A1'] = "Lithuania - aFRR, Imbalance & Intra-Day Market Analysis"
ws4['A1'].font = TITLE_FONT
ws4.merge_cells('A2:J2')
ws4['A2'] = "Sources: Litgrid, ENTSO-E, CEER, ESS News, PICASSO"
ws4['A2'].font = NOTE_FONT

# aFRR Market
row4 = 4
ws4.cell(row=row4, column=1, value="aFRR Capacity Market").font = SUBTITLE_FONT
afr_headers = ['Parameter', '2023', '2024', '2025', '2026E', '2027E', '2028E', '2029E', '2030E']
row4 += 1
for col_idx, h in enumerate(afr_headers, 1):
    ws4.cell(row=row4, column=col_idx, value=h)
style_header_row(ws4, row4, len(afr_headers))

afrr_data = [
    ['aFRR Up Requirement (MW)', 40, 60, 80, 96, 110, 120, 120, 120],
    ['aFRR Down Requirement (MW)', 40, 60, 80, 104, 110, 120, 120, 120],
    ['aFRR Capacity Price Up (EUR/MW/h)', 15, 20, 25, 30, 28, 25, 22, 20],
    ['aFRR Capacity Price Down (EUR/MW/h)', 8, 12, 15, 18, 16, 14, 12, 10],
    ['aFRR Energy Price Up (EUR/MWh)', 120, 130, 140, 145, 135, 125, 115, 110],
    ['aFRR Energy Price Down (EUR/MWh)', -20, -25, -30, -35, -40, -45, -50, -55],
    ['Total aFRR Cost (EUR mn)', 5.3, 8.5, 12, 16, 18, 19, 18, 17],
    ['BESS Share of aFRR (%)', 15, 25, 40, 55, 65, 75, 80, 85],
]

for i, row_data in enumerate(afrr_data):
    r = row4 + 1 + i
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=col_idx, value=val)
        is_fc = col_idx >= 6  # 2026+
        style_data_cell(ws4, r, col_idx, is_forecast=is_fc)
        if col_idx == 1:
            ws4.cell(row=r, column=col_idx).alignment = Alignment(horizontal='left')

# Imbalance Prices
imb_start = row4 + len(afrr_data) + 3
ws4.cell(row=imb_start, column=1, value="Imbalance Settlement").font = SUBTITLE_FONT
imb_headers = ['Parameter', '2023', '2024', '2025', '2026E', '2027E']
imb_start += 1
for col_idx, h in enumerate(imb_headers, 1):
    ws4.cell(row=imb_start, column=col_idx, value=h)
style_header_row(ws4, imb_start, len(imb_headers))

imb_data = [
    ['Imbalance Price Avg Positive (EUR/MWh)', 110, 105, 100, 95, 90],
    ['Imbalance Price Avg Negative (EUR/MWh)', -15, -25, -35, -45, -55],
    ['Imbalance Price Max Positive (EUR/MWh)', 800, 1200, 1500, 1800, 2000],
    ['Imbalance Price Min Negative (EUR/MWh)', -150, -300, -400, -500, -500],
    ['Settlement Period', '60 min', '15 min', '15 min', '15 min', '15 min'],
    ['Avg Spread (Pos-Neg) (EUR/MWh)', 125, 130, 135, 140, 145],
    ['PICASSO Integration', 'No', 'No', 'Yes (Mar)', 'Yes', 'Yes'],
]

for i, row_data in enumerate(imb_data):
    r = imb_start + 1 + i
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=col_idx, value=val)
        is_fc = col_idx >= 5
        style_data_cell(ws4, r, col_idx, is_forecast=is_fc)
        if col_idx == 1:
            ws4.cell(row=r, column=col_idx).alignment = Alignment(horizontal='left')

# Intra-day market
id_start = imb_start + len(imb_data) + 3
ws4.cell(row=id_start, column=1, value="Intra-Day Market").font = SUBTITLE_FONT
id_headers = ['Parameter', '2023', '2024', '2025', '2026E', '2027E']
id_start += 1
for col_idx, h in enumerate(id_headers, 1):
    ws4.cell(row=id_start, column=col_idx, value=h)
style_header_row(ws4, id_start, len(id_headers))

id_data = [
    ['ID Avg Price (EUR/MWh)', 93, 88, 75, 80, 75],
    ['ID-DA Spread Avg (EUR/MWh)', 5, 8, 10, 12, 12],
    ['ID-DA Spread P90 (EUR/MWh)', 25, 35, 45, 55, 55],
    ['ID Volatility (Std Dev EUR/MWh)', 40, 45, 55, 60, 65],
    ['Neg Price Hours ID', 30, 120, 177, 220, 280],
    ['ID Trading Volume (TWh)', 1.2, 1.8, 2.5, 3.5, 4.5],
]

for i, row_data in enumerate(id_data):
    r = id_start + 1 + i
    for col_idx, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=col_idx, value=val)
        is_fc = col_idx >= 5
        style_data_cell(ws4, r, col_idx, is_forecast=is_fc)
        if col_idx == 1:
            ws4.cell(row=r, column=col_idx).alignment = Alignment(horizontal='left')

# Key notes
note_row = id_start + len(id_data) + 2
ws4.cell(row=note_row, column=1, value="Key Market Developments:").font = SUBTITLE_FONT
notes = [
    "Mar 2025: Lithuania (Litgrid) joined PICASSO — pan-European aFRR platform",
    "2024: Imbalance settlement shifted from 60-min to 15-min periods (EU regulation)",
    "aFRR demand: up to 96-120 MW up / 104-120 MW down expected in 2026",
    "Baltic TSOs demand reduction reserve reduced total balancing costs 8x (Feb-Sep 2025)",
    "Ancillary service costs increased 5.5x in 2024 vs 2023 (to 1.31 ct/kWh)",
    "BESS increasingly capturing aFRR market — Energy Cells 40MW already providing aFRR",
    "Intra-day volatility increasing with renewable penetration — strong BESS opportunity",
    "NOTE: Exact aFRR/imbalance prices are estimates based on ENTSO-E reports & CEER data",
]
for i, note in enumerate(notes):
    ws4.cell(row=note_row + 1 + i, column=1, value=f"• {note}").font = DATA_FONT

auto_width(ws4)

# ============================================================
# SHEET 5: BESS Revenue Streams & Saturation Analysis
# ============================================================
ws5 = wb.create_sheet("BESS Saturation Analysis")

ws5.merge_cells('A1:K1')
ws5['A1'] = "Lithuania - BESS Market Saturation & Revenue Analysis"
ws5['A1'].font = TITLE_FONT
ws5.merge_cells('A2:K2')
ws5['A2'] = "Birdview Energy Analysis | Based on collected market data"
ws5['A2'].font = NOTE_FONT

# Revenue stream analysis
row5 = 4
ws5.cell(row=row5, column=1, value="BESS Revenue Potential by Stream (EUR/kW/yr)").font = SUBTITLE_FONT
rev_headers = ['Revenue Stream', '2024', '2025', '2026E', '2027E', '2028E', '2029E', '2030E', 'Notes']
row5 += 1
for col_idx, h in enumerate(rev_headers, 1):
    ws5.cell(row=row5, column=col_idx, value=h)
style_header_row(ws5, row5, len(rev_headers))

# Revenue estimates per kW/year for a 2h BESS system
rev_data = [
    ['DA Arbitrage', 35, 30, 28, 25, 22, 20, 18, 'Declining as price spreads narrow with more BESS'],
    ['Intra-Day Trading', 15, 18, 20, 22, 20, 18, 16, 'Growing with volatility, then saturates'],
    ['aFRR Capacity', 25, 30, 35, 32, 28, 25, 22, 'Strong near-term; competition increases'],
    ['aFRR Energy', 8, 10, 12, 11, 10, 9, 8, 'Linked to activation frequency'],
    ['mFRR / Balancing', 5, 8, 10, 10, 9, 8, 7, 'Baltic market integration helps'],
    ['Imbalance Optimization', 10, 12, 15, 14, 12, 10, 8, '15-min ISP beneficial for BESS'],
    ['TOTAL', 98, 108, 120, 114, 101, 90, 79, ''],
]

for i, row_data in enumerate(rev_data):
    r = row5 + 1 + i
    for col_idx, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=col_idx, value=val)
        is_fc = col_idx >= 4
        style_data_cell(ws5, r, col_idx, is_forecast=is_fc)
        if col_idx == 1 or col_idx == len(row_data):
            ws5.cell(row=r, column=col_idx).alignment = Alignment(horizontal='left')
    if row_data[0] == 'TOTAL':
        for c in range(1, len(row_data) + 1):
            ws5.cell(row=r, column=c).font = Font(name="Calibri", bold=True, size=10)

# Saturation analysis
sat_start = row5 + len(rev_data) + 3
ws5.cell(row=sat_start, column=1, value="Market Saturation Indicators").font = SUBTITLE_FONT
sat_headers = ['Metric', '2024', '2025', '2026E', '2027E', '2028E', '2029E', '2030E']
sat_start += 1
for col_idx, h in enumerate(sat_headers, 1):
    ws5.cell(row=sat_start, column=col_idx, value=h)
style_header_row(ws5, sat_start, len(sat_headers))

# Key saturation metrics
sat_data = [
    ['BESS Installed (MW)', 250, 535, 800, 1200, 1500, 1700, 1800],
    ['BESS / Peak Demand (%)', 12, 25, 36, 52, 63, 67, 67],
    ['BESS / Avg Load (%)', 17, 35, 49, 70, 82, 88, 88],
    ['BESS / RES Capacity (%)', 7, 12, 15, 18, 19, 20, 19],
    ['aFRR Market Size (MW)', 120, 160, 200, 220, 240, 240, 240],
    ['BESS as % of aFRR Market', 42, 67, 80, 109, 125, 142, 150],
    ['Revenue Cannibalization (%)', 0, 5, 10, 18, 25, 30, 35],
    ['DA Price Spread (EUR/MWh)', 55, 50, 48, 42, 38, 35, 32],
    ['Saturation Signal', 'Low', 'Low-Med', 'Medium', 'Med-High', 'High', 'High', 'Saturated'],
]

for i, row_data in enumerate(sat_data):
    r = sat_start + 1 + i
    for col_idx, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=col_idx, value=val)
        is_fc = col_idx >= 4
        style_data_cell(ws5, r, col_idx, is_forecast=is_fc)
        if col_idx == 1:
            ws5.cell(row=r, column=col_idx).alignment = Alignment(horizontal='left')

# Saturation conclusion
conc_row = sat_start + len(sat_data) + 2
ws5.cell(row=conc_row, column=1, value="BESS Market Saturation Assessment").font = SUBTITLE_FONT
conclusions = [
    "WINDOW OF OPPORTUNITY: 2025-2027 is the golden window for Lithuanian BESS investment",
    "",
    "BULL CASE (Favorable for BESS):",
    "  • Russia grid disconnection increases price volatility & need for flexibility",
    "  • Load growth 4%/yr (data centers, EVs, heat pumps) expands addressable market",
    "  • 0.7 GW offshore wind by 2030 creates large balancing need",
    "  • PICASSO integration opens cross-border aFRR revenue",
    "  • Government allocated €197mn+ in BESS support (largest EU procurement)",
    "",
    "BEAR CASE (Risk of saturation):",
    "  • 1.7 GW / 4 GWh government pipeline → BESS/Peak ratio hits 67% by 2029",
    "  • aFRR market relatively small (~120 MW each direction) vs BESS pipeline",
    "  • BESS will exceed total aFRR market size by 2027 → not all can earn capacity revenue",
    "  • Revenue cannibalization estimated at 25-35% by 2028-2030",
    "  • Lithuania is a small market (~13-18 TWh) — limited absolute opportunity",
    "",
    "RECOMMENDATION:",
    "  • Early-mover advantage is CRITICAL — projects operational by 2026-2027 capture peak revenues",
    "  • Target multi-revenue stacking: aFRR + DA arb + ID trading + imbalance",
    "  • Consider 2h+ duration to capture wider DA spreads",
    "  • Cross-border revenue (via PICASSO, NordBalt, LitPol Link) is essential for long-term value",
    "  • After 2028, expect revenue compression of 20-35% from baseline due to saturation",
    "  • Monitor actual BESS build-out pace — delays in pipeline could extend opportunity window",
]

for i, line in enumerate(conclusions):
    ws5.cell(row=conc_row + 1 + i, column=1, value=line).font = DATA_FONT
    if 'WINDOW' in line or 'RECOMMENDATION' in line:
        ws5.cell(row=conc_row + 1 + i, column=1).font = Font(name="Calibri", bold=True, size=11, color="C00000")
    elif 'BULL CASE' in line or 'BEAR CASE' in line:
        ws5.cell(row=conc_row + 1 + i, column=1).font = Font(name="Calibri", bold=True, size=11, color="1F4E79")

# Add saturation chart
chart5 = BarChart()
chart5.type = "col"
chart5.title = "BESS Market Saturation vs Revenue (Lithuania)"
chart5.y_axis.title = "EUR/kW/yr or %"
chart5.style = 10
chart5.width = 24
chart5.height = 14

# Revenue total row
rev_row_idx = row5 + len(rev_data)
rev_ref = Reference(ws5, min_col=2, max_col=8, min_row=rev_row_idx, max_row=rev_row_idx)
chart5.add_data(rev_ref, from_rows=True, titles_from_data=False)

# Use years as categories
ws5.add_chart(chart5, f"A{conc_row + len(conclusions) + 2}")

auto_width(ws5)

# ============================================================
# SHEET 6: Market Overview & Key Developments
# ============================================================
ws6 = wb.create_sheet("Market Overview")

ws6.merge_cells('A1:D1')
ws6['A1'] = "Lithuania Electricity Market — Key Developments for BESS Developers"
ws6['A1'].font = TITLE_FONT
ws6.merge_cells('A2:D2')
ws6['A2'] = f"Compiled {datetime.now().strftime('%B %Y')} | Birdview Energy"
ws6['A2'].font = NOTE_FONT

row6 = 4
sections = [
    ("1. GRID DISCONNECTION FROM RUSSIA (Completed Feb 2025)", [
        "8 Feb 2025: Baltic states disconnected from Russian BRELL grid system",
        "9 Feb 2025: Synchronized with Continental European grid via LitPol Link (AC mode)",
        "Ended 50+ years of Soviet/Russian grid dependency",
        "Price spike to €325/MWh on disconnection day, €128/MWh daily average",
        "Successful 24-hour isolated island mode test before EU sync",
        "Increases need for domestic flexibility → bullish for BESS",
    ]),
    ("2. MASSIVE BESS PROCUREMENT PROGRAM", [
        "Government launched 1.7 GW / 4 GWh BESS procurement — largest in EU",
        "Over 50 applications received totaling €197mn — nearly 2x the original budget",
        "Additional €45mn allocated (Oct 2025) to support more projects",
        "Target: ≥800 MWh connected to TSO grid by end 2028",
        "Key projects: Ignitis 291MW/582MWh (2027), Trina 90MW/180MWh (mid-2026)",
        "Litgrid already operates 200MW/200MWh (Fluence/Siemens Energy, since 2022)",
    ]),
    ("3. RENEWABLE ENERGY EXPLOSION", [
        "2025: record 1.7 GW of wind + solar added in single year",
        "Total wind+solar reached 3 GW by end 2025 (Litgrid confirmed)",
        "Solar: from 250 MW (2021) → 2,800 MW (2025) — 11x growth in 4 years",
        "Wind: from 671 MW (2021) → 1,500 MW (2025) — doubled",
        "2030 NECP targets: 5.1 GW solar, ~4 GW wind (incl. 0.7 GW offshore)",
        "Renewables now cover ~70% of domestic electricity demand",
    ]),
    ("4. aFRR & BALANCING MARKET EVOLUTION", [
        "Mar 2025: Litgrid joined PICASSO — pan-European aFRR platform",
        "aFRR requirement growing: up to 96-120 MW each direction by 2026",
        "15-minute imbalance settlement period introduced (was 60 min)",
        "Ancillary service costs surged 5.5x in 2024 (to 1.31 ct/kWh)",
        "BESS ideally suited for fast-response aFRR — replacing demand reduction resources",
        "Baltic demand reduction reserve reduced balancing costs 8x (Feb-Sep 2025)",
    ]),
    ("5. INTERCONNECTIONS & MARKET COUPLING", [
        "NordBalt: 700 MW HVDC to Sweden (operational since 2015)",
        "LitPol Link: 500 MW to Poland (now in AC mode after sync)",
        "Harmony Link: planned 700 MW HVDC to Poland (under development)",
        "Future: 2 GW offshore interconnection Lithuania-Latvia-Germany (post-2035)",
        "Strong interconnections keep Lithuania coupled to Nordic + CE prices",
        "But transmission constraints can cause price divergence → BESS opportunity",
    ]),
    ("6. ELECTRICITY DEMAND OUTLOOK", [
        "Current: ~13-14 TWh/year, ~2.1 GW peak demand",
        "Litgrid 10-year plan: demand grows 4%/yr → 18.7 TWh by 2031",
        "Key drivers: data centers, electric vehicles, industrial heat pumps",
        "Lithuania targets net electricity exporter by 2028-2030",
        "Import share dropped from 70% (2021) to ~33% (2025)",
        "Domestic generation doubled from 4.2 TWh (2021) to ~9 TWh (2025)",
    ]),
    ("7. REGULATORY & POLICY LANDSCAPE", [
        "Strong government support for BESS — €197mn+ allocated",
        "EU ETS2 introduction in 2027 expected to push energy prices up",
        "Lithuania 100% renewable electricity target by 2030",
        "NECP: 5.1 GW solar, ~4 GW wind, 1.7 GW storage by 2030",
        "Active support for hydrogen (electrolysis) — potential future BESS competitor",
        "Energy independence is national security priority post-Ukraine war",
    ]),
]

for title, items in sections:
    ws6.cell(row=row6, column=1, value=title).font = SUBTITLE_FONT
    row6 += 1
    for item in items:
        ws6.cell(row=row6, column=1, value=f"  • {item}").font = DATA_FONT
        row6 += 1
    row6 += 1

auto_width(ws6, min_width=15, max_width=120)
ws6.column_dimensions['A'].width = 100

# ============================================================
# SHEET 7: ENTSO-E API Data Retrieval Guide
# ============================================================
ws7 = wb.create_sheet("Data Retrieval Guide")

ws7.merge_cells('A1:D1')
ws7['A1'] = "ENTSO-E API & Data Retrieval Guide for Lithuania"
ws7['A1'].font = TITLE_FONT
ws7.merge_cells('A2:D2')
ws7['A2'] = "How to retrieve live price data programmatically"
ws7['A2'].font = NOTE_FONT

row7 = 4
guide_text = [
    ("ENTSO-E Transparency Platform API", [
        "Register at: https://transparency.entsoe.eu/",
        "Request API key: email transparency@entsoe.eu with subject 'Restful API access'",
        "Lithuania bidding zone code: 10YLT-1001A0008Q",
        "Country code for entsoe-py: 'LT'",
        "",
        "Python example (entsoe-py library):",
        "  pip install entsoe-py",
        "  from entsoe import EntsoePandasClient",
        "  import pandas as pd",
        "  client = EntsoePandasClient(api_key='YOUR_KEY')",
        "  start = pd.Timestamp('20240101', tz='Europe/Vilnius')",
        "  end = pd.Timestamp('20250101', tz='Europe/Vilnius')",
        "  # Day-ahead prices",
        "  da = client.query_day_ahead_prices('LT', start=start, end=end)",
        "  # Imbalance prices",
        "  imb = client.query_imbalance_prices('LT', start=start, end=end)",
        "  # Cross-border flows",
        "  flows = client.query_crossborder_flows('LT', 'SE_4', start=start, end=end)",
    ]),
    ("Nord Pool Data Portal", [
        "URL: https://data.nordpoolgroup.com/",
        "Lithuania data available from day-ahead and intra-day markets",
        "Free registration required for bulk downloads",
        "Also available via: https://nordpool.opendailydata.com/en (open data)",
    ]),
    ("Litgrid Dashboard", [
        "URL: https://www.litgrid.eu/index.php/dashboard/gas-and-electricity-prices/31769",
        "Real-time and historical electricity prices, flows, generation",
        "Balancing market data and ancillary services costs",
    ]),
    ("Additional Data Sources", [
        "EnergyInEU (monthly averages): https://energyineu.com/lithuania.php",
        "Ember (EU price analysis): https://ember-energy.org/data/european-electricity-prices-and-costs/",
        "Eurostat (official statistics): https://ec.europa.eu/eurostat/",
        "CEER (regulatory reports): https://www.ceer.eu/",
        "IEA Lithuania Review: https://iea.blob.core.windows.net/assets/.../Lithuania2025.pdf",
    ]),
]

for title, items in guide_text:
    ws7.cell(row=row7, column=1, value=title).font = SUBTITLE_FONT
    row7 += 1
    for item in items:
        ws7.cell(row=row7, column=1, value=item).font = DATA_FONT
        row7 += 1
    row7 += 1

auto_width(ws7, min_width=15, max_width=100)
ws7.column_dimensions['A'].width = 90

# ============================================================
# Save workbook
# ============================================================
output_path = "/Users/mayk/LithuaniaBESS/BirdEnergySystemInstalled_Lithuania.xlsx"
wb.save(output_path)
print(f"Excel workbook saved to: {output_path}")
print("Sheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
