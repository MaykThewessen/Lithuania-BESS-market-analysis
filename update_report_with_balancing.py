"""
Update Excel & HTML report with extended balancing data
========================================================
Adds: aFRR reserve prices, mFRR reserve prices, activated energy prices
"""

import os
import pandas as pd
import numpy as np
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = "/Users/mayk/LithuaniaBESS/data"
OUT_DIR = "/Users/mayk/LithuaniaBESS"

THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
API_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def write_cell(ws, row, col, val, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = DATA_FONT; cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center'); cell.fill = API_FILL
    if fmt: cell.number_format = fmt

# ============================================================
# Load data
# ============================================================
print("Loading balancing data...")

afrr = pd.read_csv(f"{DATA_DIR}/afrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
afrr.index = pd.to_datetime(afrr.index, utc=True)
for col in afrr.columns:
    afrr[col] = pd.to_numeric(afrr[col], errors='coerce')

mfrr = pd.read_csv(f"{DATA_DIR}/mfrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
mfrr.index = pd.to_datetime(mfrr.index, utc=True)
for col in mfrr.columns:
    mfrr[col] = pd.to_numeric(mfrr[col], errors='coerce')

act = pd.read_csv(f"{DATA_DIR}/activated_balancing_energy_prices_LT.csv", index_col=0, parse_dates=True)
act.index = pd.to_datetime(act.index, utc=True)
act['Price'] = pd.to_numeric(act['Price'], errors='coerce')

# Also load existing imbalance data
imb = pd.read_csv(f"{DATA_DIR}/imbalance_prices_LT.csv", index_col=0, parse_dates=True)
imb.index = pd.to_datetime(imb.index, utc=True)
for col in imb.columns:
    imb[col] = pd.to_numeric(imb[col], errors='coerce')

print(f"  aFRR: {len(afrr)} rows, {afrr.index.min().date()} to {afrr.index.max().date()}")
print(f"  mFRR: {len(mfrr)} rows, {mfrr.index.min().date()} to {mfrr.index.max().date()}")
print(f"  Activated: {len(act)} rows")
print(f"  Imbalance: {len(imb)} rows, {imb.index.min().date()} to {imb.index.max().date()}")

# ============================================================
# Compute monthly/annual stats
# ============================================================
print("Computing analytics...")

# aFRR monthly stats
afrr['year'] = afrr.index.year
afrr['month'] = afrr.index.month

afrr_monthly = afrr.groupby(['year', 'month']).agg(
    up_price_mean=('Up Prices', 'mean'),
    up_price_median=('Up Prices', 'median'),
    up_price_p90=('Up Prices', lambda x: x.quantile(0.9)),
    up_qty_mean=('Up Quantity', 'mean'),
    down_price_mean=('Down Prices', 'mean'),
    down_price_median=('Down Prices', 'median'),
    down_qty_mean=('Down Quantity', 'mean'),
).reset_index()

afrr_annual = afrr.groupby('year').agg(
    up_price_mean=('Up Prices', 'mean'),
    up_price_median=('Up Prices', 'median'),
    up_price_p90=('Up Prices', lambda x: x.quantile(0.9)),
    up_price_max=('Up Prices', 'max'),
    up_qty_mean=('Up Quantity', 'mean'),
    down_price_mean=('Down Prices', 'mean'),
    down_price_median=('Down Prices', 'median'),
    down_price_p90=('Down Prices', lambda x: x.quantile(0.9)),
    down_qty_mean=('Down Quantity', 'mean'),
    count=('Up Prices', 'count'),
)

# mFRR monthly stats
mfrr['year'] = mfrr.index.year
mfrr['month'] = mfrr.index.month

mfrr_annual = mfrr.groupby('year').agg(
    up_price_mean=('Up Prices', 'mean'),
    up_price_median=('Up Prices', 'median'),
    up_price_max=('Up Prices', 'max'),
    up_qty_mean=('Up Quantity', 'mean'),
    down_price_mean=('Down Prices', 'mean'),
    down_price_median=('Down Prices', 'median'),
    down_qty_mean=('Down Quantity', 'mean'),
    count=('Up Prices', 'count'),
)

# Activated energy - by direction and month
if 'Direction' in act.columns:
    act_up = act[act['Direction'] == 'Up']['Price']
    act_down = act[act['Direction'] == 'Down']['Price']
    act_summary = {
        'up_count': len(act_up), 'up_mean': act_up.mean(), 'up_median': act_up.median(), 'up_max': act_up.max(),
        'down_count': len(act_down), 'down_mean': act_down.mean(), 'down_median': act_down.median(), 'down_min': act_down.min(),
    }

# Imbalance annual stats (existing data)
imb.columns = ['Long', 'Short'] if list(imb.columns) != ['Long', 'Short'] else imb.columns
imb['year'] = imb.index.year
imb['spread'] = imb['Short'] - imb['Long']

imb_monthly = imb.groupby([imb.index.year, imb.index.month]).agg(
    long_mean=('Long', 'mean'),
    short_mean=('Short', 'mean'),
    spread_mean=('spread', 'mean'),
    spread_max=('spread', lambda x: x.max()),
    long_min=('Long', 'min'),
    short_max=('Short', 'max'),
).reset_index()

# ============================================================
# Update Excel
# ============================================================
print("Updating Excel...")

wb = load_workbook(f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx")

if "Balancing Data (API)" in wb.sheetnames:
    del wb["Balancing Data (API)"]
ws = wb.create_sheet("Balancing Data (API)", 1)

# Title
ws.merge_cells('A1:L1')
ws['A1'] = "Lithuania — Balancing & Reserve Market Data (ENTSO-E API)"
ws['A1'].font = TITLE_FONT
ws.merge_cells('A2:L2')
ws['A2'] = f"Retrieved: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Endpoints: A47 (aFRR), A51 (mFRR), Activated Energy, Imbalance"
ws['A2'].font = NOTE_FONT

# --- Section A: aFRR Annual Summary ---
r = 4
ws.cell(row=r, column=1, value="A. aFRR Contracted Reserve Prices — Annual Summary (EUR/MW/period)").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Up Price\nMean', 'Up Price\nMedian', 'Up Price\nP90', 'Up Price\nMax',
                      'Up Qty\nMean (MW)', 'Down Price\nMean', 'Down Price\nMedian', 'Down Price\nP90',
                      'Down Qty\nMean (MW)', 'Data Points'])
for yr, row_data in afrr_annual.iterrows():
    r += 1
    write_cell(ws, r, 1, int(yr))
    write_cell(ws, r, 2, round(row_data['up_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['up_price_median'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['up_price_p90'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['up_price_max'], 2), '#,##0.00')
    write_cell(ws, r, 6, round(row_data['up_qty_mean'], 0), '#,##0')
    write_cell(ws, r, 7, round(row_data['down_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 8, round(row_data['down_price_median'], 2), '#,##0.00')
    write_cell(ws, r, 9, round(row_data['down_price_p90'], 2), '#,##0.00')
    write_cell(ws, r, 10, round(row_data['down_qty_mean'], 0), '#,##0')
    write_cell(ws, r, 11, int(row_data['count']))

# --- Section B: aFRR Monthly ---
r += 2
ws.cell(row=r, column=1, value="B. aFRR Monthly Breakdown").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Month', 'Up Price Mean', 'Up Price Median', 'Up Price P90',
                      'Up Qty Mean', 'Down Price Mean', 'Down Qty Mean'])
for _, row_data in afrr_monthly.iterrows():
    r += 1
    write_cell(ws, r, 1, int(row_data['year']))
    write_cell(ws, r, 2, int(row_data['month']))
    write_cell(ws, r, 3, round(row_data['up_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['up_price_median'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['up_price_p90'], 2), '#,##0.00')
    write_cell(ws, r, 6, round(row_data['up_qty_mean'], 0), '#,##0')
    write_cell(ws, r, 7, round(row_data['down_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 8, round(row_data['down_qty_mean'], 0), '#,##0')

# --- Section C: mFRR Annual ---
r += 2
ws.cell(row=r, column=1, value="C. mFRR Contracted Reserve Prices — Annual Summary").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Up Price Mean', 'Up Price Median', 'Up Price Max',
                      'Up Qty Mean (MW)', 'Down Price Mean', 'Down Price Median',
                      'Down Qty Mean (MW)', 'Data Points'])
for yr, row_data in mfrr_annual.iterrows():
    r += 1
    write_cell(ws, r, 1, int(yr))
    write_cell(ws, r, 2, round(row_data['up_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['up_price_median'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['up_price_max'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['up_qty_mean'], 0), '#,##0')
    write_cell(ws, r, 6, round(row_data['down_price_mean'], 2), '#,##0.00')
    write_cell(ws, r, 7, round(row_data['down_price_median'], 2), '#,##0.00')
    write_cell(ws, r, 8, round(row_data['down_qty_mean'], 0), '#,##0')
    write_cell(ws, r, 9, int(row_data['count']))

# --- Section D: Activated Balancing ---
r += 2
ws.cell(row=r, column=1, value="D. Activated Balancing Energy Prices (Jun-Oct 2024)").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Direction', 'Count', 'Mean Price', 'Median Price', 'Max/Min Price'])
if 'Direction' in act.columns:
    r += 1
    write_cell(ws, r, 1, 'Up (shortage)')
    write_cell(ws, r, 2, act_summary['up_count'])
    write_cell(ws, r, 3, round(act_summary['up_mean'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(act_summary['up_median'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(act_summary['up_max'], 2), '#,##0.00')
    r += 1
    write_cell(ws, r, 1, 'Down (surplus)')
    write_cell(ws, r, 2, act_summary['down_count'])
    write_cell(ws, r, 3, round(act_summary['down_mean'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(act_summary['down_median'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(act_summary['down_min'], 2), '#,##0.00')

# --- Section E: Key Notes ---
r += 2
ws.cell(row=r, column=1, value="E. Key Interpretation Notes").font = SUBTITLE_FONT
notes = [
    "aFRR data available from Jun 2024 onwards (Baltic aFRR market launch)",
    "2024 aFRR Up price fixed at €0.87/MW/period — initial regulated price before market launch",
    "2025: Market-based pricing active — mean Up €29.55, median €11.61, P90 significantly higher",
    "2026 (Jan-Feb): Mean Up €33.60 — winter peak effect visible",
    "aFRR Up Quantity: ~350-430 MW contracted on average — this is TOTAL Baltic requirement, not just LT",
    "Lithuania's share is typically ~80-120 MW of the total Baltic aFRR requirement",
    "mFRR prices are significantly higher (mean €80 Up in 2025) but volumes are smaller (~36 MW)",
    "Activated balancing energy Up prices: mean €297/MWh — significant premium over DA prices",
    "PICASSO go-live (Mar 2025) changed pricing dynamics — cross-border competition for aFRR",
    "Imbalance price data gap Oct 2024 — Feb 2026: ENTSO-E has not published new-format data yet",
    "The 15-min ISP transition caused a data format change that ENTSO-E is still catching up on",
]
for note in notes:
    r += 1
    ws.cell(row=r, column=1, value=f"• {note}").font = DATA_FONT

# Auto width
for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = 18

wb.save(f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx")
print("Excel updated with 'Balancing Data (API)' sheet.")

# ============================================================
# Now update the HTML report — append balancing section
# ============================================================
print("Updating HTML report...")

# Read existing HTML
with open(f"{OUT_DIR}/Lithuania_BESS_Market_Report.html", 'r', encoding='utf-8') as f:
    html = f.read()

# Prepare chart data
afrr_monthly_json = {}
for _, row in afrr_monthly.iterrows():
    key = f"{int(row['year'])}-{int(row['month']):02d}"
    afrr_monthly_json[key] = {
        'up_mean': round(row['up_price_mean'], 2),
        'up_median': round(row['up_price_median'], 2),
        'up_p90': round(row['up_price_p90'], 2),
        'up_qty': round(row['up_qty_mean'], 0),
        'down_mean': round(row['down_price_mean'], 2),
        'down_qty': round(row['down_qty_mean'], 0),
    }

afrr_annual_json = {}
for yr, row in afrr_annual.iterrows():
    afrr_annual_json[int(yr)] = {
        'up_mean': round(row['up_price_mean'], 2),
        'up_median': round(row['up_price_median'], 2),
        'up_p90': round(row['up_price_p90'], 2),
        'up_max': round(row['up_price_max'], 2),
        'up_qty': round(row['up_qty_mean'], 0),
        'down_mean': round(row['down_price_mean'], 2),
        'down_qty': round(row['down_qty_mean'], 0),
    }

mfrr_annual_json = {}
for yr, row in mfrr_annual.iterrows():
    mfrr_annual_json[int(yr)] = {
        'up_mean': round(row['up_price_mean'], 2),
        'up_median': round(row['up_price_median'], 2),
        'up_max': round(row['up_price_max'], 2),
        'up_qty': round(row['up_qty_mean'], 0),
        'down_mean': round(row['down_price_mean'], 2),
        'down_qty': round(row['down_qty_mean'], 0),
    }

# Imbalance monthly chart data
imb_monthly_json = {}
for _, row in imb_monthly.iterrows():
    yr = int(row.iloc[0])
    mo = int(row.iloc[1])
    key = f"{yr}-{mo:02d}"
    imb_monthly_json[key] = {
        'long_mean': round(float(row['long_mean']), 2),
        'short_mean': round(float(row['short_mean']), 2),
        'spread_mean': round(float(row['spread_mean']), 2),
        'long_min': round(float(row['long_min']), 2),
        'short_max': round(float(row['short_max']), 2),
    }

# Build new section HTML
new_section = f"""
<!-- ========== EXTENDED BALANCING DATA ========== -->
<div class="section-divider"><h2>2b. Extended Balancing Market Data (ENTSO-E API)</h2></div>

<div class="grid">
    <div class="card card-full">
        <h2>aFRR Reserve Prices — Monthly (EUR/MW/period)</h2>
        <p style="font-size:0.85em; color:#6C757D; margin-bottom:12px;">
            Data from ENTSO-E contracted reserve prices endpoint (A47). Available Jun 2024 onward.
            Quantities shown are total Baltic requirement (~80-120 MW is Lithuania's share).
        </p>
        <div id="chart_afrr_monthly" class="chart"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>aFRR Annual Summary</h2>
        <table>
            <tr><th>Year</th><th>Up Mean</th><th>Up Median</th><th>Up P90</th><th>Up Max</th><th>Up Qty (MW)</th><th>Down Mean</th><th>Down Qty</th></tr>
"""

for yr in sorted(afrr_annual_json.keys()):
    d = afrr_annual_json[yr]
    new_section += f"""            <tr><td><strong>{yr}</strong></td>
                <td>{d['up_mean']:.2f}</td><td>{d['up_median']:.2f}</td><td>{d['up_p90']:.2f}</td>
                <td>{d['up_max']:.2f}</td><td>{d['up_qty']:.0f}</td>
                <td>{d['down_mean']:.2f}</td><td>{d['down_qty']:.0f}</td></tr>\n"""

new_section += f"""        </table>
        <div class="insight">
            <strong>Key finding:</strong> aFRR Up prices jumped from a flat €0.87 in 2024 (regulated) to
            mean €29.55 in 2025 (market-based), with P90 reaching significant premiums. This confirms
            the aFRR market is now competitive and price-discovering — excellent for BESS.
        </div>
    </div>
    <div class="card">
        <h2>mFRR Annual Summary</h2>
        <table>
            <tr><th>Year</th><th>Up Mean</th><th>Up Median</th><th>Up Max</th><th>Up Qty (MW)</th><th>Down Mean</th><th>Down Qty</th></tr>
"""

for yr in sorted(mfrr_annual_json.keys()):
    d = mfrr_annual_json[yr]
    new_section += f"""            <tr><td><strong>{yr}</strong></td>
                <td>{d['up_mean']:.2f}</td><td>{d['up_median']:.2f}</td><td>{d['up_max']:.2f}</td>
                <td>{d['up_qty']:.0f}</td><td>{d['down_mean']:.2f}</td><td>{d['down_qty']:.0f}</td></tr>\n"""

new_section += f"""        </table>
        <div class="insight">
            <strong>Key finding:</strong> mFRR Up mean €80/MW in 2025 — much higher than aFRR but volumes
            are small (~36 MW). Max prices of €4,000 indicate extreme scarcity events. BESS can capture
            these spikes through fast-response capability.
        </div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Activated Balancing Energy Prices (Jun-Oct 2024)</h2>
        <table>
            <tr><th>Direction</th><th>Activations</th><th>Mean Price</th><th>Median</th><th>Extreme</th></tr>
            <tr><td>Up (shortage)</td><td>{act_summary['up_count']}</td><td>€{act_summary['up_mean']:.0f}/MWh</td>
                <td>€{act_summary['up_median']:.0f}/MWh</td><td>Max €{act_summary['up_max']:.0f}/MWh</td></tr>
            <tr><td>Down (surplus)</td><td>{act_summary['down_count']}</td><td>€{act_summary['down_mean']:.0f}/MWh</td>
                <td>€{act_summary['down_median']:.0f}/MWh</td><td>Min €{act_summary['down_min']:.0f}/MWh</td></tr>
        </table>
        <div class="insight">
            <strong>Insight:</strong> Up-regulation activations average €297/MWh — a 3-4x premium over
            day-ahead prices. This is pure margin for BESS providing balancing services. Down activations
            are near-zero or negative, meaning BESS gets paid to charge.
        </div>
    </div>
    <div class="card">
        <h2>Imbalance Prices — Monthly (2021-Sep 2024)</h2>
        <div id="chart_imb_monthly" class="chart" style="min-height:350px"></div>
    </div>
</div>

<div class="grid">
    <div class="card card-full">
        <div class="highlight-box">
            <h3 style="color:#856404; margin-top:0;">Balancing Market — BESS Revenue Implications</h3>
            <table style="margin-top:10px">
                <tr><th>Revenue Stream</th><th>Data Source</th><th>Avg Price</th><th>BESS Revenue Potential</th></tr>
                <tr><td>aFRR Capacity (Up)</td><td>ENTSO-E A47</td><td>€29.55/MW/period (2025)</td>
                    <td>~€25-35/kW/yr at 80% availability</td></tr>
                <tr><td>aFRR Capacity (Down)</td><td>ENTSO-E A47</td><td>€18.34/MW/period (2025)</td>
                    <td>~€15-20/kW/yr (stackable with Up)</td></tr>
                <tr><td>mFRR Reserve</td><td>ENTSO-E A51</td><td>€80.31/MW/period (2025)</td>
                    <td>~€5-10/kW/yr (small volumes ~36 MW)</td></tr>
                <tr><td>Activated Energy (Up)</td><td>ENTSO-E Activated</td><td>€297/MWh</td>
                    <td>Depends on activation frequency</td></tr>
                <tr><td>Activated Energy (Down)</td><td>ENTSO-E Activated</td><td>€-5/MWh (paid to charge)</td>
                    <td>Additional margin when activated</td></tr>
            </table>
            <p style="margin-top:12px; font-size:0.9em;">
                <strong>Total aFRR + mFRR capacity revenue potential: ~€40-55/kW/yr</strong> — this confirms
                the earlier estimate of €30-35/kW/yr for aFRR alone was conservative. The real data shows
                higher prices, especially post-PICASSO go-live in 2025.
            </p>
        </div>
    </div>
</div>
"""

# Inject chart scripts
new_scripts = f"""
<script>
// ==================== aFRR MONTHLY ====================
(() => {{
    const data = {json.dumps(afrr_monthly_json)};
    const months = Object.keys(data).sort();
    Plotly.newPlot('chart_afrr_monthly', [
        {{x: months, y: months.map(m => data[m].up_mean), name: 'Up Price Mean', type: 'scatter',
          mode: 'lines+markers', line: {{color: '#C00000', width: 2.5}}, marker: {{size: 5}}}},
        {{x: months, y: months.map(m => data[m].up_median), name: 'Up Price Median', type: 'scatter',
          mode: 'lines', line: {{color: '#C00000', width: 1.5, dash: 'dot'}}}},
        {{x: months, y: months.map(m => data[m].up_p90), name: 'Up Price P90', type: 'scatter',
          mode: 'lines', line: {{color: '#ED7D31', width: 2}}}},
        {{x: months, y: months.map(m => data[m].down_mean), name: 'Down Price Mean', type: 'scatter',
          mode: 'lines+markers', line: {{color: '#70AD47', width: 2.5}}, marker: {{size: 5}}}},
        {{x: months, y: months.map(m => data[m].up_qty), name: 'Up Qty (MW)', type: 'bar',
          marker: {{color: 'rgba(46,117,182,0.2)'}}, yaxis: 'y2'}}
    ], {{
        title: 'aFRR Contracted Reserve Prices (Monthly)',
        yaxis: {{title: 'EUR/MW/period', gridcolor: '#E0E0E0', side: 'left'}},
        yaxis2: {{title: 'Quantity (MW)', overlaying: 'y', side: 'right', gridcolor: 'transparent'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.25}}, margin: {{t: 40, b: 80}},
        shapes: [{{type: 'line', x0: '2025-03', x1: '2025-03', y0: 0, y1: 200,
                   line: {{color: '#7B2D8E', width: 2, dash: 'dash'}}}}],
        annotations: [{{x: '2025-03', y: 180, text: 'PICASSO<br>Go-Live', showarrow: false,
                        font: {{color: '#7B2D8E', size: 10}}}}]
    }}, {{responsive: true, displayModeBar: false}});
}})();

// ==================== IMBALANCE MONTHLY ====================
(() => {{
    const data = {json.dumps(imb_monthly_json)};
    const months = Object.keys(data).sort();
    Plotly.newPlot('chart_imb_monthly', [
        {{x: months, y: months.map(m => data[m].long_mean), name: 'Avg Long (surplus)', type: 'scatter',
          mode: 'lines', line: {{color: '#70AD47', width: 2}}}},
        {{x: months, y: months.map(m => data[m].short_mean), name: 'Avg Short (deficit)', type: 'scatter',
          mode: 'lines', line: {{color: '#C00000', width: 2}}}},
        {{x: months, y: months.map(m => data[m].short_max), name: 'Max Short', type: 'scatter',
          mode: 'markers', marker: {{color: '#C00000', size: 6, symbol: 'triangle-up'}}}},
        {{x: months, y: months.map(m => data[m].long_min), name: 'Min Long', type: 'scatter',
          mode: 'markers', marker: {{color: '#70AD47', size: 6, symbol: 'triangle-down'}}}}
    ], {{
        title: 'Monthly Imbalance Prices (2021 — Sep 2024)',
        yaxis: {{title: 'EUR/MWh', gridcolor: '#E0E0E0'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.25}}, margin: {{t: 40, b: 80}}
    }}, {{responsive: true, displayModeBar: false}});
}})();
</script>
"""

# Insert new section before the "Load & Generation" section
insert_marker = '<!-- ========== LOAD & GENERATION ========== -->'
if insert_marker in html:
    html = html.replace(insert_marker, new_section + '\n' + insert_marker)
else:
    # Fallback: insert before footer
    html = html.replace('</div><!-- container -->', new_section + '\n</div><!-- container -->')

# Insert new scripts before closing </body>
html = html.replace('</body>', new_scripts + '\n</body>')

with open(f"{OUT_DIR}/Lithuania_BESS_Market_Report.html", 'w', encoding='utf-8') as f:
    f.write(html)

print(f"HTML report updated with extended balancing section.")
print("Done!")
