"""
Add BESS Revenue Analysis Section to HTML Report & Excel
=========================================================
Computes revenue by duration (1h/2h/4h) for each market:
  DA Arbitrage, aFRR, FCR, mFRR, Imbalance, Multi-Market Combined
Integrates LinkedIn market intelligence on Baltic balancing markets.
"""

import os
import pandas as pd
import numpy as np
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

DATA_DIR = "/Users/mayk/LithuaniaBESS/data"
OUT_DIR = "/Users/mayk/LithuaniaBESS"

# Styling constants
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
API_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=10, color="1F4E79")

# BESS parameters
RT_EFF = 0.88                # round-trip efficiency
DA_CAPTURE = 0.85            # realistic vs perfect-foresight DA arbitrage
DURATIONS = [1, 2, 4]        # hours
PERIODS_PER_YEAR = 35_040    # 15-min ISPs in a year

# aFRR/mFRR availability by duration (limited by SoC)
AFRR_AVAIL = {1: 0.65, 2: 0.80, 4: 0.90}
# FCR availability by duration (good for all BESS, SoC managed)
FCR_AVAIL = {1: 0.90, 2: 0.92, 4: 0.95}
# mFRR availability
MFRR_AVAIL = {1: 0.70, 2: 0.82, 4: 0.90}

# FCR price estimate (no ENTSO-E data; from LinkedIn market intelligence)
# Baltic FCR launched Feb 5 2025; first day €145/MW/h (highest in Europe)
# Conservative annual average estimate considering market maturation
FCR_PRICE_PER_HOUR = {
    2024: 0,      # FCR market didn't exist in Baltics
    2025: 30,     # EUR/MW/h average (high due to new market, scarcity)
    2026: 22,     # declining as more capacity enters
}

# Multi-market time allocation weights
# BESS optimally splits between highest-value markets
MULTI_MARKET_ALLOC = {
    'aFRR':      0.40,   # highest value capacity market
    'FCR':       0.20,   # second priority (symmetric, good for BESS)
    'mFRR':      0.05,   # low volume, occasional
    'DA':        0.25,   # remaining time for arbitrage
    'Imbalance': 0.10,   # opportunistic
}


# ============================================================
# Load data
# ============================================================
print("Loading market data...")

da = pd.read_csv(f"{DATA_DIR}/da_prices_LT.csv", index_col=0, parse_dates=True)
da.columns = ['price']
da.index = pd.to_datetime(da.index, utc=True)
da['price'] = pd.to_numeric(da['price'], errors='coerce')

afrr = pd.read_csv(f"{DATA_DIR}/afrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
afrr.index = pd.to_datetime(afrr.index, utc=True)
for col in afrr.columns:
    afrr[col] = pd.to_numeric(afrr[col], errors='coerce')

mfrr = pd.read_csv(f"{DATA_DIR}/mfrr_reserve_prices_LT.csv", index_col=0, parse_dates=True)
mfrr.index = pd.to_datetime(mfrr.index, utc=True)
for col in mfrr.columns:
    mfrr[col] = pd.to_numeric(mfrr[col], errors='coerce')

imb = pd.read_csv(f"{DATA_DIR}/imbalance_prices_LT.csv", index_col=0, parse_dates=True)
imb.index = pd.to_datetime(imb.index, utc=True)
for col in imb.columns:
    imb[col] = pd.to_numeric(imb[col], errors='coerce')

print(f"  DA prices: {len(da)} rows ({da.index.min()} to {da.index.max()})")
print(f"  aFRR: {len(afrr)} rows ({afrr.index.min()} to {afrr.index.max()})")
print(f"  mFRR: {len(mfrr)} rows ({mfrr.index.min()} to {mfrr.index.max()})")
print(f"  Imbalance: {len(imb)} rows ({imb.index.min()} to {imb.index.max()})")


# ============================================================
# 1. DA Arbitrage Revenue
# ============================================================
def compute_da_revenue(da_prices, year, duration):
    """Perfect-foresight DA arbitrage: buy N cheapest, sell N most expensive hours per day."""
    mask = da_prices.index.year == year
    if mask.sum() == 0:
        return 0
    yearly = da_prices.loc[mask, 'price'].dropna()
    daily = yearly.groupby(yearly.index.date)
    total_rev = 0
    days = 0
    for _, group in daily:
        if len(group) < 24:
            continue
        prices = group.values
        sorted_p = np.sort(prices)
        buy = sorted_p[:duration].sum()
        sell = sorted_p[-duration:].sum()
        rev = sell * np.sqrt(RT_EFF) - buy / np.sqrt(RT_EFF)
        total_rev += max(0, rev)
        days += 1
    if days == 0:
        return 0
    return (total_rev / days * 365) * DA_CAPTURE


# ============================================================
# 2. aFRR Capacity Revenue
# ============================================================
def compute_afrr_revenue(afrr_data, year, duration):
    """aFRR capacity revenue from contracted reserve prices.

    ENTSO-E reports prices per MW per ISP (15-min period).
    BESS provides Up capacity (discharge) and Down capacity (charge),
    alternating to manage SoC. Revenue = sum of all ISP prices × availability.
    """
    mask = afrr_data.index.year == year
    if mask.sum() == 0:
        return 0
    yr_data = afrr_data.loc[mask]
    actual_periods = len(yr_data)
    ann_factor = PERIODS_PER_YEAR / actual_periods if actual_periods > 0 else 0

    up_sum = yr_data['Up Prices'].sum()
    down_sum = yr_data['Down Prices'].sum()

    # BESS alternates Up/Down to manage SoC: ~50% time each direction
    # Revenue from both directions with availability factor
    avail = AFRR_AVAIL[duration]
    revenue = (0.5 * up_sum + 0.5 * down_sum) * avail * ann_factor
    return revenue


# ============================================================
# 3. FCR Revenue (estimated from LinkedIn market data)
# ============================================================
def compute_fcr_revenue(year, duration):
    """FCR revenue estimate based on Baltic market intelligence.

    Baltic FCR market launched Feb 5, 2025. First day: €145/MW/h.
    FCR is ideal for BESS: symmetric, fast-response, limited energy needed.
    """
    price = FCR_PRICE_PER_HOUR.get(year, 15)  # default declining
    avail = FCR_AVAIL[duration]
    hours = 8760
    if year == 2025:
        hours = 8760 * (11 / 12)  # market started in February
    return price * hours * avail


# ============================================================
# 4. mFRR Revenue
# ============================================================
def compute_mfrr_revenue(mfrr_data, year, duration):
    """mFRR capacity revenue. Similar to aFRR but lower volumes and sporadic."""
    mask = mfrr_data.index.year == year
    if mask.sum() == 0:
        return 0
    yr_data = mfrr_data.loc[mask]
    actual_periods = len(yr_data)
    ann_factor = PERIODS_PER_YEAR / actual_periods if actual_periods > 0 else 0

    up_sum = yr_data['Up Prices'].sum()
    down_sum = yr_data['Down Prices'].sum()

    avail = MFRR_AVAIL[duration]
    revenue = (0.5 * up_sum + 0.5 * down_sum) * avail * ann_factor
    return revenue


# ============================================================
# 5. Imbalance Trading Revenue
# ============================================================
def compute_imbalance_revenue(da_prices, imb_prices, year, duration):
    """Revenue from DA vs imbalance price spread trading.

    BESS positions based on expected imbalance direction:
    - System short (imb > DA): discharge at premium
    - System long (imb < DA): charge at discount
    Revenue = |imbalance - DA| captured during best N hours/day.
    """
    mask_da = da_prices.index.year == year
    mask_imb = imb_prices.index.year == year
    if mask_da.sum() == 0 or mask_imb.sum() == 0:
        return 0

    common = da_prices.loc[mask_da].index.intersection(imb_prices.loc[mask_imb].index)
    if len(common) == 0:
        return 0

    da_c = da_prices.loc[common, 'price']
    imb_c = imb_prices.loc[common, 'Short']  # single imbalance price
    spread = (imb_c - da_c).abs()

    daily = spread.groupby(spread.index.date)
    total = 0
    days = 0
    for _, group in daily:
        best_n = group.nlargest(duration)
        total += best_n.sum() * RT_EFF
        days += 1
    if days == 0:
        return 0
    return total / days * 365 * DA_CAPTURE  # same capture rate as DA


# ============================================================
# Compute all revenues
# ============================================================
print("\nComputing BESS revenue estimates...")

YEARS = [2024, 2025]
results = {}  # {(year, market, duration): revenue}

for year in YEARS:
    for dur in DURATIONS:
        # DA Arbitrage
        da_rev = compute_da_revenue(da, year, dur)
        results[(year, 'DA Arbitrage', dur)] = da_rev

        # aFRR
        afrr_rev = compute_afrr_revenue(afrr, year, dur)
        results[(year, 'aFRR', dur)] = afrr_rev

        # FCR
        fcr_rev = compute_fcr_revenue(year, dur)
        results[(year, 'FCR', dur)] = fcr_rev

        # mFRR
        mfrr_rev = compute_mfrr_revenue(mfrr, year, dur)
        results[(year, 'mFRR', dur)] = mfrr_rev

        # Imbalance
        imb_rev = compute_imbalance_revenue(da, imb, year, dur)
        results[(year, 'Imbalance', dur)] = imb_rev

        # Multi-Market Combined
        combined = (
            MULTI_MARKET_ALLOC['aFRR'] * afrr_rev / AFRR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['FCR'] * fcr_rev / FCR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['mFRR'] * mfrr_rev / MFRR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['DA'] * da_rev / DA_CAPTURE +
            MULTI_MARKET_ALLOC['Imbalance'] * imb_rev / DA_CAPTURE
        )
        results[(year, 'Multi-Market Combined', dur)] = combined

MARKETS = ['DA Arbitrage', 'aFRR', 'FCR', 'mFRR', 'Imbalance', 'Multi-Market Combined']

# Print summary
for year in YEARS:
    print(f"\n{'='*70}")
    print(f"  BESS Revenue Estimates {year} (EUR/MW/year)")
    print(f"{'='*70}")
    print(f"  {'Market':<25} {'1h BESS':>12} {'2h BESS':>12} {'4h BESS':>12}")
    print(f"  {'-'*25} {'-'*12} {'-'*12} {'-'*12}")
    for mkt in MARKETS:
        vals = [results.get((year, mkt, d), 0) for d in DURATIONS]
        print(f"  {mkt:<25} {vals[0]:>12,.0f} {vals[1]:>12,.0f} {vals[2]:>12,.0f}")

# ============================================================
# Forward projections (2026-2030) with revenue compression
# ============================================================
print("\nComputing forward projections with BESS saturation effects...")

# Revenue compression factors (declining from extraordinary 2025 levels)
# Based on: 454 MW installed today → 1,200+ MW by 2027 → 2,000+ MW by 2030
COMPRESSION = {
    2024: 1.0,
    2025: 1.0,   # actual data
    2026: 0.65,   # significant new BESS coming online
    2027: 0.45,   # market approaching saturation
    2028: 0.35,   # saturated for aFRR; DA still valuable
    2029: 0.30,   # mature market
    2030: 0.28,   # steady state
}

# DA arbitrage compresses less than balancing (structural price spreads persist)
DA_COMPRESSION = {
    2024: 1.0,
    2025: 1.0,
    2026: 0.85,
    2027: 0.70,
    2028: 0.60,
    2029: 0.55,
    2030: 0.50,
}

# FCR grows then stabilizes (market matures)
FCR_PRICE_FORECAST = {
    2024: 0, 2025: 30, 2026: 22, 2027: 18, 2028: 15, 2029: 12, 2030: 10
}

proj_results = {}
for year in range(2025, 2031):
    comp = COMPRESSION.get(year, 0.28)
    da_comp = DA_COMPRESSION.get(year, 0.50)
    for dur in DURATIONS:
        # Use 2025 as base year for balancing markets
        base_afrr = results.get((2025, 'aFRR', dur), 0)
        base_mfrr = results.get((2025, 'mFRR', dur), 0)
        base_da = results.get((2025, 'DA Arbitrage', dur), 0)
        base_imb = results.get((2024, 'Imbalance', dur), 0)  # 2024 has full data

        proj_results[(year, 'DA Arbitrage', dur)] = base_da * da_comp
        proj_results[(year, 'aFRR', dur)] = base_afrr * comp
        proj_results[(year, 'FCR', dur)] = compute_fcr_revenue(year, dur)
        proj_results[(year, 'mFRR', dur)] = base_mfrr * comp
        proj_results[(year, 'Imbalance', dur)] = base_imb * da_comp

        # Multi-market combined (optimized stacking)
        combined = (
            MULTI_MARKET_ALLOC['aFRR'] * base_afrr * comp / AFRR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['FCR'] * compute_fcr_revenue(year, dur) / FCR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['mFRR'] * base_mfrr * comp / MFRR_AVAIL[dur] +
            MULTI_MARKET_ALLOC['DA'] * base_da * da_comp / DA_CAPTURE +
            MULTI_MARKET_ALLOC['Imbalance'] * base_imb * da_comp / DA_CAPTURE
        )
        proj_results[(year, 'Multi-Market Combined', dur)] = combined

# Print projections
for year in range(2025, 2031):
    print(f"\n  {year} (compression: balancing={COMPRESSION.get(year, 0.28):.0%}, DA={DA_COMPRESSION.get(year, 0.50):.0%}):")
    for dur in DURATIONS:
        comb = proj_results.get((year, 'Multi-Market Combined', dur), 0)
        print(f"    {dur}h BESS Combined: {comb:,.0f} EUR/MW/year")


# ============================================================
# LinkedIn Market Intelligence Summary
# ============================================================
LINKEDIN_INSIGHTS = [
    {
        'source': 'Balancing Services OÜ',
        'date': 'Feb 5, 2025',
        'key': 'Baltic balancing capacity procurement launched. Latvia: mFRR UP 55 MW, mFRR DOWN 55 MW, '
               'FCR 29 MW. Lithuania joined for mFRR DOWN 440 MW. FCR price hit €145/MW/h — '
               'one of the highest in Europe. mFRR DOWN: €20–30/MW/h. Peak mFRR UP: €65/MW/h (Latvia) '
               'vs €0.7/MW/h (Lithuania) showing extreme early-market volatility.',
    },
    {
        'source': 'Fusebox Energy',
        'date': 'Mar 2025',
        'key': 'Baltic frequency reserves showed extreme price swings from €9,976/MWh to -€4,473/MWh. '
               'A single small Latvian bidder caused market distortions. Regulators investigating '
               'potential manipulation. Highlights need for more storage capacity and market design reform.',
    },
    {
        'source': 'Energy Lead',
        'date': 'Aug 6, 2025',
        'key': 'Nord Pool day-ahead prices hit -€3.06/MWh across all three Baltic states. '
               'Ancillary services saw -€11,999/MWh in Latvia and Lithuania — '
               'unprecedented volatility showing extreme market dynamics.',
    },
    {
        'source': 'Danny Zaitsev (Zada)',
        'date': 'Jun 2025',
        'key': 'Baltic region remains an "energy island" with prices disconnected from Europe. '
               '"A battery doing simple DA trading will be earning in excess of EUR 800/MW." '
               'Lithuania approved 4 GWh of battery storage. Opportunity "won\'t last forever — '
               'once new storage is built, prices will settle down."',
    },
    {
        'source': 'Elektrum Eesti',
        'date': 'Feb 2025',
        'key': 'Estlink 2 cable failure caused 65–71% Baltic electricity price surge. '
               'Estonia local production covered only 64% of consumption. '
               'Demonstrates BESS upside from interconnection failures driving scarcity pricing.',
    },
    {
        'source': 'Litgrid',
        'date': 'Jun 2025',
        'key': 'Lithuania\'s first commercial BESS connected: 53.6 MW / 107.3 MWh by UAB Vėjo Galia '
               'in Kaišiadorys district. Total Lithuania BESS: 453.9 MW / 461.5 MWh. '
               'At least 9 more storage projects planned for next year.',
    },
    {
        'source': 'Estonian Consumer Impact',
        'date': 'Jan 2026',
        'key': 'Post-BRELL disconnection costs: Estonian consumers now pay €3.70/MWh reserve capacity fee '
               'plus ~€2/month "quiet sleep tax". Demonstrates structural cost of energy independence '
               'creating sustained revenue pool for flexibility providers.',
    },
    {
        'source': 'VPPA/Spread Analysis',
        'date': '2025',
        'key': 'Baltic States daily DA price spread: €177 (comparable to Austria €187–222). '
               'Analysis of 20 million ENTSO-E rows across 38 European regions confirms '
               'Baltic grid decoupling driving flexibility demand. Volatility is "fuel for battery profits."',
    },
    {
        'source': 'Grid Disconnection',
        'date': 'Feb 7–9, 2025',
        'key': 'Estonia, Latvia, Lithuania disconnected from Russian BRELL grid and joined Continental '
               'European Network. Total project cost: €1.2 billion (75% EU funded). '
               'TSOs deployed synchronous condensers and BESS for frequency management.',
    },
]


# ============================================================
# UPDATE EXCEL
# ============================================================
print("\n\nUpdating Excel workbook...")

XLSX_PATH = f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx"

def write_header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def write_data_cell(ws, row, col, val, fmt=None, fill=None, font=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    cell.fill = fill or API_FILL
    if fmt:
        cell.number_format = fmt

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)

wb = load_workbook(XLSX_PATH)

# Remove existing sheet if present
if "BESS Revenue Analysis" in wb.sheetnames:
    del wb["BESS Revenue Analysis"]

ws = wb.create_sheet("BESS Revenue Analysis", 2)  # 3rd position

row = 1
ws.cell(row=row, column=1, value="BESS Revenue Analysis by Duration & Market").font = TITLE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1
ws.cell(row=row, column=1,
        value="Based on actual ENTSO-E data (2024-2025) with forward projections (2026-2030)").font = NOTE_FONT
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 2

# ---- Table 1: Actual Revenue 2025 ----
ws.cell(row=row, column=1,
        value="Table 1: BESS Revenue 2025 Actual — EUR/MW/year").font = SUBTITLE_FONT
row += 1
write_header_row(ws, row, ['Revenue Stream', '1h BESS', '2h BESS', '4h BESS',
                            '1h EUR/kW/yr', '2h EUR/kW/yr', '4h EUR/kW/yr'])
row += 1
for mkt in MARKETS:
    vals = [results.get((2025, mkt, d), 0) for d in DURATIONS]
    fill = TOTAL_FILL if mkt == 'Multi-Market Combined' else API_FILL
    font = TOTAL_FONT if mkt == 'Multi-Market Combined' else DATA_FONT
    write_data_cell(ws, row, 1, mkt, fill=fill, font=font)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    for i, v in enumerate(vals):
        write_data_cell(ws, row, 2 + i, round(v), fmt='#,##0', fill=fill, font=font)
        write_data_cell(ws, row, 5 + i, round(v / 1000, 1), fmt='#,##0.0', fill=fill, font=font)
    row += 1

row += 1

# ---- Table 2: Actual Revenue 2024 ----
ws.cell(row=row, column=1,
        value="Table 2: BESS Revenue 2024 Actual — EUR/MW/year").font = SUBTITLE_FONT
row += 1
write_header_row(ws, row, ['Revenue Stream', '1h BESS', '2h BESS', '4h BESS',
                            '1h EUR/kW/yr', '2h EUR/kW/yr', '4h EUR/kW/yr'])
row += 1
for mkt in MARKETS:
    vals = [results.get((2024, mkt, d), 0) for d in DURATIONS]
    fill = TOTAL_FILL if mkt == 'Multi-Market Combined' else API_FILL
    font = TOTAL_FONT if mkt == 'Multi-Market Combined' else DATA_FONT
    write_data_cell(ws, row, 1, mkt, fill=fill, font=font)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
    for i, v in enumerate(vals):
        write_data_cell(ws, row, 2 + i, round(v), fmt='#,##0', fill=fill, font=font)
        write_data_cell(ws, row, 5 + i, round(v / 1000, 1), fmt='#,##0.0', fill=fill, font=font)
    row += 1

row += 1

# ---- Table 3: Forward Projections (Multi-Market Combined) ----
ws.cell(row=row, column=1,
        value="Table 3: Multi-Market Combined Revenue Projection — EUR/MW/year").font = SUBTITLE_FONT
row += 1
proj_table_start = row
write_header_row(ws, row, ['Year', '1h BESS', '2h BESS', '4h BESS',
                            'Balancing Compression', 'DA Compression', 'BESS Installed (MW)'])
row += 1
bess_installed = {2025: 454, 2026: 700, 2027: 1200, 2028: 1500, 2029: 1800, 2030: 2000}
for year in range(2025, 2031):
    vals = [proj_results.get((year, 'Multi-Market Combined', d), 0) for d in DURATIONS]
    write_data_cell(ws, row, 1, year)
    for i, v in enumerate(vals):
        write_data_cell(ws, row, 2 + i, round(v), fmt='#,##0')
    write_data_cell(ws, row, 5, f"{COMPRESSION.get(year, 0.28):.0%}")
    write_data_cell(ws, row, 6, f"{DA_COMPRESSION.get(year, 0.50):.0%}")
    write_data_cell(ws, row, 7, bess_installed.get(year, ''))
    row += 1

# Add chart
chart = BarChart()
chart.type = "col"
chart.grouping = "clustered"
chart.title = "Multi-Market Combined Revenue Projection (EUR/MW/year)"
chart.y_axis.title = "EUR/MW/year"
chart.x_axis.title = "Year"
chart.style = 10
chart.width = 22
chart.height = 13

cats = Reference(ws, min_col=1, min_row=proj_table_start + 1,
                 max_row=proj_table_start + 6)
for col_idx, label in [(2, '1h BESS'), (3, '2h BESS'), (4, '4h BESS')]:
    data_ref = Reference(ws, min_col=col_idx, min_row=proj_table_start,
                         max_row=proj_table_start + 6)
    chart.add_data(data_ref, titles_from_data=True)

chart.set_categories(cats)
chart.shape = 4
ws.add_chart(chart, f"A{row + 1}")
row += 18

# ---- Table 4: Revenue Breakdown by Market (2025) ----
ws.cell(row=row, column=1,
        value="Table 4: Revenue Breakdown by Market Segment 2025 — EUR/MW/year").font = SUBTITLE_FONT
row += 1
breakdown_start = row
write_header_row(ws, row, ['Duration', 'DA Arbitrage', 'aFRR', 'FCR',
                            'mFRR', 'Imbalance', 'Combined'])
row += 1
for dur in DURATIONS:
    write_data_cell(ws, row, 1, f"{dur}h BESS")
    for i, mkt in enumerate(MARKETS):
        v = results.get((2025, mkt, dur), 0)
        fill = TOTAL_FILL if mkt == 'Multi-Market Combined' else API_FILL
        font = TOTAL_FONT if mkt == 'Multi-Market Combined' else DATA_FONT
        write_data_cell(ws, row, 2 + i, round(v), fmt='#,##0', fill=fill, font=font)
    row += 1

# Add stacked bar chart for 2025 breakdown
chart2 = BarChart()
chart2.type = "col"
chart2.grouping = "stacked"
chart2.title = "Revenue by Market Segment 2025 (EUR/MW/year)"
chart2.y_axis.title = "EUR/MW/year"
chart2.style = 10
chart2.width = 22
chart2.height = 13

cats2 = Reference(ws, min_col=1, min_row=breakdown_start + 1,
                  max_row=breakdown_start + 3)
for col_idx in range(2, 7):  # DA through Imbalance (not Combined)
    data_ref = Reference(ws, min_col=col_idx, min_row=breakdown_start,
                         max_row=breakdown_start + 3)
    chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats2)
ws.add_chart(chart2, f"A{row + 1}")
row += 18

# ---- LinkedIn Market Intelligence ----
ws.cell(row=row, column=1,
        value="Baltic Market Intelligence (LinkedIn Sources)").font = SUBTITLE_FONT
row += 1
write_header_row(ws, row, ['Source', 'Date', 'Key Insight'])
row += 1
for insight in LINKEDIN_INSIGHTS:
    write_data_cell(ws, row, 1, insight['source'])
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', wrap_text=True)
    write_data_cell(ws, row, 2, insight['date'])
    write_data_cell(ws, row, 3, insight['key'])
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', wrap_text=True)
    row += 1

row += 1

# ---- Methodology Notes ----
ws.cell(row=row, column=1, value="Methodology & Assumptions").font = SUBTITLE_FONT
row += 1
notes = [
    "DA Arbitrage: Perfect-foresight upper bound (buy N cheapest, sell N most expensive hours/day) × 85% capture rate.",
    f"Round-trip efficiency: {RT_EFF:.0%}. Accounts for charge/discharge losses in Li-ion BESS.",
    "aFRR: ENTSO-E contracted reserve prices (process type A47). Prices per MW per 15-min ISP. BESS alternates Up/Down.",
    f"aFRR availability by duration: 1h={AFRR_AVAIL[1]:.0%}, 2h={AFRR_AVAIL[2]:.0%}, 4h={AFRR_AVAIL[4]:.0%} (SoC constraints).",
    "FCR: Estimated from Baltic market data (launched Feb 2025). No ENTSO-E data available for Lithuania FCR.",
    f"FCR price estimate 2025: €{FCR_PRICE_PER_HOUR[2025]}/MW/h average. First-day clearing: €145/MW/h (LinkedIn source).",
    "mFRR: ENTSO-E process type A51. Lower volumes than aFRR. Sporadic procurement.",
    "Imbalance: |Imbalance price - DA price| spread captured during best N hours/day. Data through Sep 2024.",
    "Multi-Market Combined: Weighted time allocation across markets (aFRR 40%, FCR 20%, DA 25%, mFRR 5%, Imbalance 10%).",
    "Forward projections apply revenue compression from BESS saturation: 454 MW (2025) → 2,000+ MW (2030).",
    "2025 revenues reflect extraordinary post-BRELL disconnection conditions. NOT sustainable long-term.",
    "aFRR revenues are exceptionally high due to scarcity; expect 55-70% decline by 2027 as capacity enters.",
    "All figures are gross revenue before opex, degradation, and financing costs.",
]
for note in notes:
    ws.cell(row=row, column=1, value=note).font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

auto_width(ws)
# Override width for key columns
ws.column_dimensions['A'].width = 30
ws.column_dimensions['C'].width = 60 if row > 50 else 15

wb.save(XLSX_PATH)
print(f"  Excel updated: added 'BESS Revenue Analysis' sheet")


# ============================================================
# UPDATE HTML REPORT
# ============================================================
print("\nUpdating HTML report...")

HTML_PATH = f"{OUT_DIR}/Lithuania_BESS_Market_Report.html"
with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

# Build the revenue data for Plotly
# Chart 1: Stacked bar - 2025 revenue by market segment
chart1_markets = ['DA Arbitrage', 'aFRR', 'FCR', 'mFRR', 'Imbalance']
chart1_data = {}
for mkt in chart1_markets:
    chart1_data[mkt] = [results.get((2025, mkt, d), 0) for d in DURATIONS]

# Chart 2: Multi-market combined projection over years
chart2_years = list(range(2025, 2031))
chart2_data = {}
for dur in DURATIONS:
    chart2_data[f'{dur}h BESS'] = [
        proj_results.get((y, 'Multi-Market Combined', dur), 0)
        for y in chart2_years
    ]

# Build table HTML for 2025 actual revenues
table_2025_rows = ""
for mkt in MARKETS:
    vals = [results.get((2025, mkt, d), 0) for d in DURATIONS]
    is_total = mkt == 'Multi-Market Combined'
    cls = ' class="total-row"' if is_total else ''
    table_2025_rows += f"<tr{cls}>"
    table_2025_rows += f"<td style='text-align:left; font-weight:{'bold' if is_total else 'normal'}'>{mkt}</td>"
    for v in vals:
        table_2025_rows += f"<td style='text-align:right; font-weight:{'bold' if is_total else 'normal'}'>{v:,.0f}</td>"
    table_2025_rows += "</tr>"

# Build table for 2024
table_2024_rows = ""
for mkt in MARKETS:
    vals = [results.get((2024, mkt, d), 0) for d in DURATIONS]
    is_total = mkt == 'Multi-Market Combined'
    cls = ' class="total-row"' if is_total else ''
    table_2024_rows += f"<tr{cls}>"
    table_2024_rows += f"<td style='text-align:left; font-weight:{'bold' if is_total else 'normal'}'>{mkt}</td>"
    for v in vals:
        table_2024_rows += f"<td style='text-align:right; font-weight:{'bold' if is_total else 'normal'}'>{v:,.0f}</td>"
    table_2024_rows += "</tr>"

# Projection table
proj_table_rows = ""
for year in range(2025, 2031):
    vals = [proj_results.get((year, 'Multi-Market Combined', d), 0) for d in DURATIONS]
    comp = COMPRESSION.get(year, 0.28)
    installed = bess_installed.get(year, '')
    proj_table_rows += f"<tr>"
    proj_table_rows += f"<td>{year}</td>"
    for v in vals:
        proj_table_rows += f"<td style='text-align:right'>{v:,.0f}</td>"
    proj_table_rows += f"<td style='text-align:right'>{comp:.0%}</td>"
    proj_table_rows += f"<td style='text-align:right'>{installed}</td>"
    proj_table_rows += "</tr>"

# LinkedIn insights HTML
linkedin_html = ""
for insight in LINKEDIN_INSIGHTS:
    linkedin_html += f"""
    <div style="background:#f8f9fa; border-left:4px solid #1F4E79; padding:12px 16px; margin:8px 0; border-radius:4px;">
        <strong>{insight['source']}</strong> <span style="color:#666; font-size:0.85em">({insight['date']})</span>
        <p style="margin:6px 0 0 0; color:#333; font-size:0.92em">{insight['key']}</p>
    </div>"""

new_section = f"""
<!-- ========== BESS REVENUE ANALYSIS ========== -->
<div style="margin-top:50px;">
    <h2 style="color:#1F4E79; border-bottom:3px solid #1F4E79; padding-bottom:10px;">
        3. BESS Revenue Analysis by Duration & Market Segment
    </h2>

    <p style="color:#555; font-size:0.95em;">
        Revenue estimates computed from actual ENTSO-E market data for Lithuania.
        Covers Day-Ahead arbitrage, aFRR/mFRR capacity payments, FCR (estimated), and imbalance trading.
        Forward projections account for revenue compression as BESS capacity grows from 454 MW (2025) toward 2,000+ MW (2030).
    </p>

    <!-- Stacked Bar Chart: 2025 Revenue by Market -->
    <h3 style="color:#2E75B6; margin-top:30px;">Revenue by Market Segment — 2025 Actual</h3>
    <div id="revChart2025" style="width:100%; height:500px;"></div>

    <!-- Revenue Table 2025 -->
    <h3 style="color:#2E75B6; margin-top:30px;">Detailed Revenue Breakdown 2025 (EUR/MW/year)</h3>
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px; text-align:left">Revenue Stream</th>
                <th style="padding:10px">1h BESS</th>
                <th style="padding:10px">2h BESS</th>
                <th style="padding:10px">4h BESS</th>
            </tr>
        </thead>
        <tbody>
            {table_2025_rows}
        </tbody>
    </table>
    </div>

    <!-- Revenue Table 2024 -->
    <h3 style="color:#2E75B6; margin-top:30px;">Revenue Comparison: 2024 (Pre-Desynchronization) (EUR/MW/year)</h3>
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px; text-align:left">Revenue Stream</th>
                <th style="padding:10px">1h BESS</th>
                <th style="padding:10px">2h BESS</th>
                <th style="padding:10px">4h BESS</th>
            </tr>
        </thead>
        <tbody>
            {table_2024_rows}
        </tbody>
    </table>
    </div>

    <!-- Projection Chart -->
    <h3 style="color:#2E75B6; margin-top:40px;">Multi-Market Combined Revenue Projection (2025–2030)</h3>
    <div id="revProjection" style="width:100%; height:500px;"></div>

    <!-- Projection Table -->
    <h3 style="color:#2E75B6; margin-top:30px;">Revenue Projection with BESS Saturation (EUR/MW/year)</h3>
    <div style="overflow-x:auto;">
    <table style="border-collapse:collapse; width:100%; font-size:0.9em;">
        <thead>
            <tr style="background:#1F4E79; color:white;">
                <th style="padding:10px">Year</th>
                <th style="padding:10px">1h BESS</th>
                <th style="padding:10px">2h BESS</th>
                <th style="padding:10px">4h BESS</th>
                <th style="padding:10px">Compression</th>
                <th style="padding:10px">BESS MW</th>
            </tr>
        </thead>
        <tbody>
            {proj_table_rows}
        </tbody>
    </table>
    </div>

    <!-- Baltic Market Intelligence -->
    <h3 style="color:#2E75B6; margin-top:40px;">Baltic Market Intelligence (Industry Sources)</h3>
    <p style="color:#555; font-size:0.9em;">
        Key insights from industry participants on the Baltic balancing and BESS market dynamics:
    </p>
    {linkedin_html}

    <!-- Key Takeaways -->
    <h3 style="color:#2E75B6; margin-top:40px;">Key Takeaways for BESS Investors</h3>
    <div style="background:linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%); color:white; padding:24px; border-radius:8px; margin:16px 0;">
        <div style="display:grid; grid-template-columns:1fr 1fr; gap:20px;">
            <div>
                <h4 style="color:#BDD7EE; margin:0 0 10px 0;">Extraordinary 2025 Revenues</h4>
                <p style="margin:0; font-size:0.92em;">Post-BRELL disconnection created scarcity: aFRR Up prices averaged
                EUR 29.56/MW per 15-min period (vs EUR 0.87 regulated in 2024). Baltic FCR hit EUR 145/MW/h on launch day.
                Extreme spikes to EUR 9,976/MWh in frequency reserves.</p>
            </div>
            <div>
                <h4 style="color:#BDD7EE; margin:0 0 10px 0;">Revenue Compression Ahead</h4>
                <p style="margin:0; font-size:0.92em;">454 MW installed today growing to 2,000+ MW by 2030.
                "Opportunity won't last forever — once new storage is built, prices will settle down" (industry source).
                Early movers capture 3-5x returns vs late entrants.</p>
            </div>
            <div>
                <h4 style="color:#BDD7EE; margin:0 0 10px 0;">Duration Matters</h4>
                <p style="margin:0; font-size:0.92em;">4h BESS earns ~3x the DA arbitrage revenue of 1h BESS.
                But for aFRR/FCR capacity markets, duration premium is smaller (1.4x).
                Optimal choice depends on market mix and entry timing.</p>
            </div>
            <div>
                <h4 style="color:#BDD7EE; margin:0 0 10px 0;">Structural Drivers Persist</h4>
                <p style="margin:0; font-size:0.92em;">Baltic "energy island" status drives price disconnection from Europe.
                Estlink 2 failures cause 65-71% price surges. Growing renewables (1.7 GW added 2025)
                increase flexibility demand. Reserve capacity fees (EUR 3.70/MWh) create sustained revenue pool.</p>
            </div>
        </div>
    </div>

    <!-- Methodology -->
    <details style="margin-top:20px;">
        <summary style="cursor:pointer; color:#2E75B6; font-weight:bold;">Methodology & Assumptions</summary>
        <div style="padding:12px; background:#f8f9fa; border-radius:4px; margin-top:8px; font-size:0.88em; color:#555;">
            <ul>
                <li><strong>DA Arbitrage:</strong> Perfect-foresight upper bound (buy N cheapest, sell N most expensive hours/day) x 85% capture rate for realistic operation.</li>
                <li><strong>Round-trip efficiency:</strong> {RT_EFF:.0%} (typical Li-ion NMC/LFP).</li>
                <li><strong>aFRR:</strong> ENTSO-E contracted reserve prices (process type A47). Prices per MW per 15-min ISP. BESS alternates Up (50%) / Down (50%) to manage SoC.</li>
                <li><strong>aFRR availability:</strong> 1h: {AFRR_AVAIL[1]:.0%}, 2h: {AFRR_AVAIL[2]:.0%}, 4h: {AFRR_AVAIL[4]:.0%} — limited by state-of-charge constraints.</li>
                <li><strong>FCR:</strong> Estimated from Baltic market data (launched Feb 2025, no ENTSO-E data). Average EUR {FCR_PRICE_PER_HOUR[2025]}/MW/h (first-day clearing: EUR 145/MW/h).</li>
                <li><strong>mFRR:</strong> ENTSO-E process type A51. Lower volumes than aFRR, sporadic procurement.</li>
                <li><strong>Imbalance:</strong> |Imbalance price - DA price| spread x best N hours/day. Data available through Sep 2024.</li>
                <li><strong>Multi-Market Combined:</strong> Weighted time allocation — aFRR {MULTI_MARKET_ALLOC['aFRR']:.0%}, FCR {MULTI_MARKET_ALLOC['FCR']:.0%}, DA {MULTI_MARKET_ALLOC['DA']:.0%}, mFRR {MULTI_MARKET_ALLOC['mFRR']:.0%}, Imbalance {MULTI_MARKET_ALLOC['Imbalance']:.0%}.</li>
                <li><strong>Forward projections:</strong> Revenue compression from 454 MW (2025) to 2,000 MW (2030). Balancing markets compress faster than DA arbitrage.</li>
                <li><strong>All figures are gross revenue</strong> before opex (~EUR 5-8/kW/yr), degradation (2-3%/yr capacity fade), and financing costs.</li>
            </ul>
        </div>
    </details>
</div>
"""

# JavaScript for Plotly charts
new_scripts = f"""
<script>
(function() {{
    // Chart 1: Stacked bar - 2025 revenue by market
    var durations = ['1h BESS', '2h BESS', '4h BESS'];
    var traces1 = [
        {{name: 'DA Arbitrage', x: durations, y: {json.dumps(chart1_data['DA Arbitrage'])}, type: 'bar',
          marker: {{color: '#2E75B6'}}}},
        {{name: 'aFRR', x: durations, y: {json.dumps(chart1_data['aFRR'])}, type: 'bar',
          marker: {{color: '#ED7D31'}}}},
        {{name: 'FCR', x: durations, y: {json.dumps(chart1_data['FCR'])}, type: 'bar',
          marker: {{color: '#A5A5A5'}}}},
        {{name: 'mFRR', x: durations, y: {json.dumps(chart1_data['mFRR'])}, type: 'bar',
          marker: {{color: '#FFC000'}}}},
        {{name: 'Imbalance', x: durations, y: {json.dumps(chart1_data['Imbalance'])}, type: 'bar',
          marker: {{color: '#70AD47'}}}}
    ];
    var layout1 = {{
        barmode: 'stack',
        title: 'BESS Revenue by Market Segment — 2025 Actual (EUR/MW/year)',
        yaxis: {{title: 'EUR/MW/year', tickformat: ','}},
        legend: {{orientation: 'h', y: -0.15}},
        margin: {{t: 50, b: 80}},
        plot_bgcolor: '#fafafa'
    }};
    Plotly.newPlot('revChart2025', traces1, layout1, {{responsive: true, displayModeBar: false}});

    // Chart 2: Multi-market projection
    var years = {json.dumps(chart2_years)};
    var yearLabels = years.map(String);
    var traces2 = [
        {{name: '1h BESS', x: yearLabels, y: {json.dumps([round(v) for v in chart2_data['1h BESS']])},
          type: 'bar', marker: {{color: '#2E75B6'}}}},
        {{name: '2h BESS', x: yearLabels, y: {json.dumps([round(v) for v in chart2_data['2h BESS']])},
          type: 'bar', marker: {{color: '#ED7D31'}}}},
        {{name: '4h BESS', x: yearLabels, y: {json.dumps([round(v) for v in chart2_data['4h BESS']])},
          type: 'bar', marker: {{color: '#70AD47'}}}}
    ];
    var layout2 = {{
        barmode: 'group',
        title: 'Multi-Market Combined Revenue Projection (EUR/MW/year)',
        yaxis: {{title: 'EUR/MW/year', tickformat: ','}},
        legend: {{orientation: 'h', y: -0.15}},
        margin: {{t: 50, b: 80}},
        plot_bgcolor: '#fafafa',
        annotations: [{{
            x: '2025', y: {round(max(proj_results.get((2025, 'Multi-Market Combined', d), 0) for d in DURATIONS))},
            text: 'Post-BRELL<br>scarcity peak',
            showarrow: true, arrowhead: 2, ax: 40, ay: -40,
            font: {{size: 11, color: '#c00'}}
        }}, {{
            x: '2027', y: {round(max(proj_results.get((2027, 'Multi-Market Combined', d), 0) for d in DURATIONS))},
            text: '~1,200 MW BESS<br>installed',
            showarrow: true, arrowhead: 2, ax: 40, ay: -40,
            font: {{size: 11, color: '#666'}}
        }}]
    }};
    Plotly.newPlot('revProjection', traces2, layout2, {{responsive: true, displayModeBar: false}});
}})();
</script>
"""

# Inject section — place before the balancing section or at end
insert_marker = '<!-- ========== EXTENDED BALANCING DATA ========== -->'
if insert_marker not in html:
    insert_marker = '<!-- ========== LOAD & GENERATION ========== -->'

if insert_marker in html:
    html = html.replace(insert_marker, new_section + '\n' + insert_marker)
else:
    html = html.replace('</div><!-- container -->', new_section + '\n</div><!-- container -->')

# Insert scripts before closing </body>
html = html.replace('</body>', new_scripts + '\n</body>')

# Add CSS for table styling
table_css = """
<style>
.total-row { background: #D6DCE4 !important; font-weight: bold; }
table td, table th { padding: 8px 12px; border: 1px solid #ddd; }
table tbody tr:nth-child(even) { background: #f2f7fb; }
table tbody tr:hover { background: #e2efda; }
details summary:hover { text-decoration: underline; }
</style>
"""
html = html.replace('</head>', table_css + '\n</head>')

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"  HTML report updated with BESS Revenue Analysis section")

print("\n" + "=" * 60)
print("SUMMARY")
print("=" * 60)
print(f"  Excel: Added 'BESS Revenue Analysis' sheet to {XLSX_PATH}")
print(f"  HTML:  Added Section 3 (Revenue Analysis) to {HTML_PATH}")
print(f"\n  2025 Multi-Market Combined Revenue (EUR/MW/year):")
for dur in DURATIONS:
    v = results.get((2025, 'Multi-Market Combined', dur), 0)
    print(f"    {dur}h BESS: {v:>12,.0f} EUR/MW/year  ({v/1000:.1f} EUR/kW/year)")
print(f"\n  2030 Projected (with saturation):")
for dur in DURATIONS:
    v = proj_results.get((2030, 'Multi-Market Combined', dur), 0)
    print(f"    {dur}h BESS: {v:>12,.0f} EUR/MW/year  ({v/1000:.1f} EUR/kW/year)")
print("\nDone!")
