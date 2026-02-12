"""
Lithuania BESS Market Analysis — Full Report Builder
=====================================================
Reads real ENTSO-E data, updates Excel, and generates HTML report
with interactive Plotly charts.
"""

import pandas as pd
import numpy as np
import json
import os
from datetime import datetime

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(_SCRIPT_DIR, "data")
OUT_DIR = _SCRIPT_DIR

# ============================================================
# 1. Load all ENTSO-E data
# ============================================================
print("Loading ENTSO-E data...")

da_prices = pd.read_csv(f"{DATA_DIR}/da_prices_LT.csv", index_col=0, parse_dates=True)
da_prices.index = pd.to_datetime(da_prices.index, utc=True)
da_prices.columns = ['price'] if len(da_prices.columns) == 1 else da_prices.columns
if isinstance(da_prices, pd.DataFrame):
    da_prices = da_prices.iloc[:, 0]
da_prices.name = 'price'

imb_prices = pd.read_csv(f"{DATA_DIR}/imbalance_prices_LT.csv", index_col=0, parse_dates=True)
imb_prices.index = pd.to_datetime(imb_prices.index, utc=True)

load_data = pd.read_csv(f"{DATA_DIR}/actual_load_LT.csv", index_col=0, parse_dates=True)
load_data.index = pd.to_datetime(load_data.index, utc=True)
load_data.columns = ['load_mw']

gen_data = pd.read_csv(f"{DATA_DIR}/generation_by_type_LT.csv", index_col=0, parse_dates=True, low_memory=False)
gen_data.index = pd.to_datetime(gen_data.index, utc=True)
# Convert all columns to numeric
for col in gen_data.columns:
    gen_data[col] = pd.to_numeric(gen_data[col], errors='coerce')

# Cross-border flows
flows = {}
for f in ['flow_SE_4_to_LT', 'flow_LT_to_SE_4', 'flow_PL_to_LT', 'flow_LT_to_PL',
          'flow_LV_to_LT', 'flow_LT_to_LV']:
    df = pd.read_csv(f"{DATA_DIR}/{f}.csv", index_col=0, parse_dates=True)
    df.index = pd.to_datetime(df.index, utc=True)
    if isinstance(df, pd.DataFrame):
        df = df.iloc[:, 0]
    flows[f] = df

print(f"  DA prices: {len(da_prices)} rows, {da_prices.index.min()} to {da_prices.index.max()}")
print(f"  Imbalance: {len(imb_prices)} rows")
print(f"  Load: {len(load_data)} rows")
print(f"  Generation: {len(gen_data)} rows, cols={list(gen_data.columns)}")

# ============================================================
# 2. Compute analytics
# ============================================================
print("\nComputing analytics...")

# --- Day-Ahead Price Analysis ---
da_df = da_prices.to_frame('price')
da_df['year'] = da_df.index.year
da_df['month'] = da_df.index.month
da_df['hour'] = da_df.index.hour

# Monthly averages
da_monthly = da_df.groupby(['year', 'month'])['price'].agg(['mean', 'min', 'max', 'std', 'count']).reset_index()
da_monthly.columns = ['year', 'month', 'avg_price', 'min_price', 'max_price', 'std_price', 'count']

# Annual averages
da_annual = da_df.groupby('year')['price'].agg(['mean', 'min', 'max', 'std', 'count']).reset_index()
da_annual.columns = ['year', 'avg_price', 'min_price', 'max_price', 'std_price', 'count']

# Negative price hours per year
neg_hours = da_df[da_df['price'] < 0].groupby('year').size().reset_index(name='neg_hours')

# Hourly price profile by year
hourly_profile = da_df.groupby(['year', 'hour'])['price'].mean().unstack('hour')

# DA price spread (daily max - min)
da_df['date'] = da_df.index.date
daily_spread = da_df.groupby(['date', 'year'])['price'].agg(['max', 'min'])
daily_spread['spread'] = daily_spread['max'] - daily_spread['min']
annual_spread = daily_spread.reset_index().groupby('year')['spread'].agg(['mean', 'median', 'quantile']).reset_index() if False else None
# simpler approach
spread_by_year = daily_spread.reset_index().groupby('year').agg(
    avg_spread=('spread', 'mean'),
    median_spread=('spread', 'median'),
    p90_spread=('spread', lambda x: x.quantile(0.9)),
    p10_spread=('spread', lambda x: x.quantile(0.1)),
).reset_index()

# --- Imbalance Price Analysis ---
imb_prices.columns = ['long_price', 'short_price'] if len(imb_prices.columns) == 2 else imb_prices.columns
imb_df = imb_prices.copy()
imb_df['year'] = imb_df.index.year
imb_df['spread'] = imb_df.iloc[:, 1] - imb_df.iloc[:, 0]  # short - long

imb_annual = imb_df.groupby('year').agg(
    avg_long=('long_price', 'mean') if 'long_price' in imb_df.columns else (imb_df.columns[0], 'mean'),
    avg_short=('short_price', 'mean') if 'short_price' in imb_df.columns else (imb_df.columns[1], 'mean'),
    max_short=('short_price', 'max') if 'short_price' in imb_df.columns else (imb_df.columns[1], 'max'),
    min_long=('long_price', 'min') if 'long_price' in imb_df.columns else (imb_df.columns[0], 'min'),
    avg_spread=('spread', 'mean'),
)

# --- Load Analysis ---
load_df = load_data.copy()
load_df['year'] = load_df.index.year

load_annual = load_df.groupby('year').agg(
    avg_load=('load_mw', 'mean'),
    min_load=('load_mw', 'min'),
    max_load=('load_mw', 'max'),
    std_load=('load_mw', 'std'),
    count=('load_mw', 'count'),
).reset_index()

# Estimate annual TWh from average MW
# For sub-hourly data, we need to account for resolution
load_annual['hours'] = load_annual['count'] * 0.25  # 15-min resolution for 2024+ data
# Fix for 2021-2023 which are hourly
for idx, row in load_annual.iterrows():
    if row['year'] <= 2023:
        load_annual.at[idx, 'hours'] = row['count']  # hourly data

load_annual['twh'] = load_annual['avg_load'] * load_annual['hours'] / 1e6

# Monthly load profile
load_df['month'] = load_df.index.month
monthly_load = load_df.groupby(['year', 'month'])['load_mw'].agg(['mean', 'min', 'max']).reset_index()

# --- Generation Analysis ---
gen_df = gen_data.copy()
gen_df['year'] = gen_df.index.year
gen_df['total'] = gen_df.sum(axis=1, numeric_only=True)

# Annual generation by type (MWh -> TWh)
gen_annual = gen_df.groupby('year').sum(numeric_only=True)
# Convert to TWh (sum of MW * 0.25h for 15min data, or * 1h for hourly)
# We need to detect resolution per year
for year in gen_annual.index:
    year_data = gen_df[gen_df['year'] == year]
    if len(year_data) > 10000:  # 15-min data
        gen_annual.loc[year] = gen_annual.loc[year] * 0.25 / 1e6  # MW * 0.25h / 1e6 = TWh
    else:
        gen_annual.loc[year] = gen_annual.loc[year] / 1e6  # MW * 1h / 1e6 = TWh

# Rename columns for clarity
gen_cols_map = {
    'Biomass': 'Biomass',
    'Fossil Gas': 'Gas',
    'Hydro Pumped Storage': 'Pumped Storage (Gen)',
    'Hydro Pumped Storage.1': 'Pumped Storage (Pump)',
    'Hydro Run-of-river and poundage': 'Hydro',
    'Other': 'Other',
    'Solar': 'Solar',
    'Waste': 'Waste',
    'Wind Onshore': 'Wind',
    'Energy storage': 'Battery Storage',
}
gen_annual_clean = gen_annual.rename(columns=gen_cols_map)

# --- Cross-border Flow Analysis ---
net_imports = {}
for year in range(2021, 2027):
    total_import = 0
    total_export = 0
    for key, series in flows.items():
        year_data = series[series.index.year == year]
        if len(year_data) == 0:
            continue
        # Detect resolution
        factor = 0.25 if len(year_data) > 10000 else 1.0
        twh = year_data.sum() * factor / 1e6
        if '_to_LT' in key:
            total_import += twh
        else:
            total_export += twh
    net_imports[year] = {
        'import_twh': total_import,
        'export_twh': total_export,
        'net_import_twh': total_import - total_export,
    }

flow_summary = pd.DataFrame(net_imports).T
flow_summary.index.name = 'year'

print("  Analytics computed.")

# ============================================================
# 3. Update Excel with real data
# ============================================================
print("\nUpdating Excel workbook with real ENTSO-E data...")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="1F4E79")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")
API_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

wb = load_workbook(f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx")

# Add a new sheet with real ENTSO-E data
if "ENTSO-E Real Data" in wb.sheetnames:
    del wb["ENTSO-E Real Data"]
ws = wb.create_sheet("ENTSO-E Real Data", 0)

# Title
ws.merge_cells('A1:O1')
ws['A1'] = "Lithuania — ENTSO-E Real Market Data (API Retrieved)"
ws['A1'].font = TITLE_FONT
ws.merge_cells('A2:O2')
ws['A2'] = f"Retrieved: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Source: ENTSO-E Transparency Platform API"
ws['A2'].font = NOTE_FONT

def write_header(ws, row, headers):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

def write_cell(ws, row, col, val, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center')
    cell.fill = API_FILL
    if fmt:
        cell.number_format = fmt

# --- Section A: Annual DA Price Summary ---
r = 4
ws.cell(row=r, column=1, value="A. Day-Ahead Price Summary (EUR/MWh)").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Avg Price', 'Min Price', 'Max Price', 'Std Dev', 'Neg Hours', 'Hours Total'])
for _, row_data in da_annual.iterrows():
    r += 1
    year = int(row_data['year'])
    write_cell(ws, r, 1, year)
    write_cell(ws, r, 2, round(row_data['avg_price'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['min_price'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['max_price'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['std_price'], 2), '#,##0.00')
    nh = neg_hours[neg_hours['year'] == year]
    write_cell(ws, r, 6, int(nh['neg_hours'].values[0]) if len(nh) > 0 else 0)
    write_cell(ws, r, 7, int(row_data['count']))

# --- Section B: DA Price Spread (Daily Max-Min) ---
r += 2
ws.cell(row=r, column=1, value="B. Daily DA Price Spread — BESS Arbitrage Potential (EUR/MWh)").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Avg Spread', 'Median Spread', 'P90 Spread', 'P10 Spread'])
for _, row_data in spread_by_year.iterrows():
    r += 1
    write_cell(ws, r, 1, int(row_data['year']))
    write_cell(ws, r, 2, round(row_data['avg_spread'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['median_spread'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['p90_spread'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['p10_spread'], 2), '#,##0.00')

# --- Section C: Imbalance Price Summary ---
r += 2
ws.cell(row=r, column=1, value="C. Imbalance Price Summary (EUR/MWh) — 2021-Sep 2024").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Avg Long', 'Avg Short', 'Min Long', 'Max Short', 'Avg Spread'])
for year, row_data in imb_annual.iterrows():
    r += 1
    write_cell(ws, r, 1, int(year))
    write_cell(ws, r, 2, round(row_data['avg_long'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['avg_short'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['min_long'], 2), '#,##0.00')
    write_cell(ws, r, 5, round(row_data['max_short'], 2), '#,##0.00')
    write_cell(ws, r, 6, round(row_data['avg_spread'], 2), '#,##0.00')

# --- Section D: Load Summary ---
r += 2
ws.cell(row=r, column=1, value="D. Electricity Load Summary").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Avg Load (MW)', 'Min Load (MW)', 'Max Load (MW)', 'Std (MW)', 'Est. TWh'])
for _, row_data in load_annual.iterrows():
    r += 1
    write_cell(ws, r, 1, int(row_data['year']))
    write_cell(ws, r, 2, round(row_data['avg_load'], 0), '#,##0')
    write_cell(ws, r, 3, round(row_data['min_load'], 0), '#,##0')
    write_cell(ws, r, 4, round(row_data['max_load'], 0), '#,##0')
    write_cell(ws, r, 5, round(row_data['std_load'], 0), '#,##0')
    write_cell(ws, r, 6, round(row_data['twh'], 2), '#,##0.00')

# --- Section E: Generation by Type ---
r += 2
ws.cell(row=r, column=1, value="E. Generation by Type (TWh)").font = SUBTITLE_FONT
r += 1
gen_headers = ['Year'] + list(gen_annual_clean.columns)
write_header(ws, r, gen_headers[:12])  # limit columns
for year, row_data in gen_annual_clean.iterrows():
    r += 1
    write_cell(ws, r, 1, int(year))
    for i, col in enumerate(gen_annual_clean.columns[:11], 2):
        val = row_data[col]
        write_cell(ws, r, i, round(val, 3) if pd.notna(val) else 0, '#,##0.000')

# --- Section F: Cross-border Flows ---
r += 2
ws.cell(row=r, column=1, value="F. Cross-border Flows (TWh)").font = SUBTITLE_FONT
r += 1
write_header(ws, r, ['Year', 'Total Import', 'Total Export', 'Net Import'])
for year, row_data in flow_summary.iterrows():
    r += 1
    write_cell(ws, r, 1, int(year))
    write_cell(ws, r, 2, round(row_data['import_twh'], 2), '#,##0.00')
    write_cell(ws, r, 3, round(row_data['export_twh'], 2), '#,##0.00')
    write_cell(ws, r, 4, round(row_data['net_import_twh'], 2), '#,##0.00')

# Auto-width
for col in ws.columns:
    from openpyxl.utils import get_column_letter
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = 16

wb.save(f"{OUT_DIR}/BirdEnergySystemInstalled_Lithuania.xlsx")
print("  Excel updated with ENTSO-E Real Data sheet.")

# ============================================================
# 4. Build HTML Report
# ============================================================
print("\nBuilding HTML report...")

# Prepare chart data as JSON for Plotly
def series_to_json(series, name='value'):
    """Convert pandas series to JSON-friendly format."""
    return {
        'x': [str(x) for x in series.index],
        'y': [float(v) if pd.notna(v) else None for v in series.values],
        'name': name
    }

# Monthly DA prices for heatmap
monthly_prices = {}
for year in sorted(da_monthly['year'].unique()):
    yr_data = da_monthly[da_monthly['year'] == year]
    monthly_prices[int(year)] = {int(r['month']): round(r['avg_price'], 1) for _, r in yr_data.iterrows()}

# Hourly profiles
hourly_data = {}
for year in hourly_profile.index:
    hourly_data[int(year)] = {int(h): round(float(v), 1) for h, v in hourly_profile.loc[year].items() if pd.notna(v)}

# Annual generation by type
gen_chart_data = {}
for col in gen_annual_clean.columns:
    vals = gen_annual_clean[col].fillna(0)
    gen_chart_data[col] = {int(y): round(float(v), 3) for y, v in vals.items()}

# Load annual
load_chart = load_annual.set_index('year')[['avg_load', 'min_load', 'max_load']].to_dict('index')

# Flow data
flow_chart = flow_summary.to_dict('index')

# Spread data
spread_chart = spread_by_year.set_index('year').to_dict('index')

# Imbalance annual
imb_chart = imb_annual.to_dict('index')

# ============================================================
# Installed capacity data (from original analysis)
# ============================================================
years_cap = [2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030, 2031]
wind_cumul = [671, 946, 1050, 1200, 1500, 1900, 2400, 3000, 3500, 4000, 4500]
pv_cumul = [250, 570, 1410, 2280, 2800, 3500, 4200, 4800, 5100, 5500, 5800]
bess_power = [0, 200, 200, 250, 535, 800, 1200, 1500, 1700, 1800, 1900]
bess_energy = [0, 200, 200, 300, 700, 1200, 2000, 3000, 4000, 4200, 4500]
fossil_mw = [1800, 1800, 1800, 1800, 1800, 1600, 1500, 1400, 1200, 1000, 900]

# BESS saturation
bess_revenue = [0, 0, 0, 75, 572, 387, 272, 220, 194, 182, 170]
bess_saturation_pct = [0, 10, 10, 12, 25, 36, 52, 63, 67, 67, 68]

# Negative price hours real
neg_hours_real = neg_hours.set_index('year')['neg_hours'].to_dict()

# ============================================================
# Generate HTML
# ============================================================
html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lithuania BESS Market Analysis — Birdview Energy</title>
    <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
    <style>
        :root {{
            --primary: #1F4E79;
            --accent: #2E75B6;
            --green: #70AD47;
            --solar: #FFC000;
            --red: #C00000;
            --bg: #F8F9FA;
            --card: #FFFFFF;
            --text: #2C3E50;
            --muted: #6C757D;
            --border: #DEE2E6;
        }}
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}
        header {{
            background: linear-gradient(135deg, var(--primary), var(--accent));
            color: white; padding: 40px 0; margin-bottom: 30px;
        }}
        header h1 {{ font-size: 2.2em; font-weight: 700; }}
        header p {{ font-size: 1.1em; opacity: 0.9; margin-top: 8px; }}
        .badge {{ display: inline-block; background: rgba(255,255,255,0.2); padding: 4px 12px;
                  border-radius: 20px; font-size: 0.85em; margin-top: 10px; }}
        .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(320px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .card {{
            background: var(--card); border-radius: 12px; padding: 24px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08); border: 1px solid var(--border);
        }}
        .card-full {{ grid-column: 1 / -1; }}
        .card h2 {{ color: var(--primary); font-size: 1.3em; margin-bottom: 16px;
                    padding-bottom: 8px; border-bottom: 2px solid var(--accent); }}
        .card h3 {{ color: var(--accent); font-size: 1.1em; margin: 16px 0 8px; }}
        .stat-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 12px; }}
        .stat {{
            background: var(--bg); border-radius: 8px; padding: 16px; text-align: center;
        }}
        .stat .value {{ font-size: 1.8em; font-weight: 700; color: var(--primary); }}
        .stat .label {{ font-size: 0.85em; color: var(--muted); margin-top: 4px; }}
        .stat.green .value {{ color: var(--green); }}
        .stat.solar .value {{ color: #E6A800; }}
        .stat.red .value {{ color: var(--red); }}
        table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 0.9em; }}
        th {{ background: var(--primary); color: white; padding: 10px 8px; text-align: center;
             font-weight: 600; }}
        td {{ padding: 8px; text-align: center; border-bottom: 1px solid var(--border); }}
        tr:nth-child(even) {{ background: #F8F9FA; }}
        tr:hover {{ background: #EBF5FB; }}
        .forecast {{ background: #FFF8E1 !important; }}
        .forecast:hover {{ background: #FFF3CD !important; }}
        .tag {{ display: inline-block; padding: 2px 10px; border-radius: 12px; font-size: 0.8em;
                font-weight: 600; }}
        .tag-green {{ background: #D4EDDA; color: #155724; }}
        .tag-yellow {{ background: #FFF3CD; color: #856404; }}
        .tag-orange {{ background: #FFE0B2; color: #E65100; }}
        .tag-red {{ background: #F8D7DA; color: #721C24; }}
        .chart {{ width: 100%; min-height: 400px; }}
        .insight {{
            background: linear-gradient(135deg, #EBF5FB, #D4E6F1);
            border-left: 4px solid var(--accent);
            padding: 16px 20px; border-radius: 0 8px 8px 0; margin: 12px 0;
        }}
        .insight strong {{ color: var(--primary); }}
        .bullet {{ margin: 6px 0; padding-left: 20px; position: relative; }}
        .bullet::before {{ content: "\\2022"; position: absolute; left: 4px; color: var(--accent); font-weight: bold; }}
        .section-divider {{
            margin: 40px 0 30px; padding: 20px 0; border-top: 3px solid var(--accent);
        }}
        .section-divider h2 {{ font-size: 1.6em; color: var(--primary); }}
        footer {{ text-align: center; padding: 30px; color: var(--muted); font-size: 0.9em;
                  border-top: 1px solid var(--border); margin-top: 40px; }}
        .highlight-box {{
            background: linear-gradient(135deg, #FFF8E1, #FFFDE7);
            border: 2px solid var(--solar);
            border-radius: 12px; padding: 20px; margin: 16px 0;
        }}
        .warning-box {{
            background: linear-gradient(135deg, #FCE4EC, #F8BBD0);
            border: 2px solid var(--red);
            border-radius: 12px; padding: 20px; margin: 16px 0;
        }}
        @media (max-width: 768px) {{
            header h1 {{ font-size: 1.5em; }}
            .grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>

<header>
    <div class="container">
        <h1>Lithuania BESS Market Analysis</h1>
        <p>Comprehensive battery energy storage market assessment — historical data & forward outlook</p>
        <span class="badge">Data: ENTSO-E API | {datetime.now().strftime('%d %B %Y')}</span>
        <span class="badge">Period: 2021 — 2031</span>
        <span class="badge">Birdview Energy</span>
    </div>
</header>

<div class="container">

<!-- ========== KEY METRICS ========== -->
<div class="grid">
    <div class="card">
        <div class="stat-grid">
            <div class="stat">
                <div class="value">{round(da_annual[da_annual['year']==2025]['avg_price'].values[0], 1) if 2025 in da_annual['year'].values else 'N/A'}</div>
                <div class="label">2025 Avg DA Price<br>EUR/MWh</div>
            </div>
            <div class="stat red">
                <div class="value">{round(da_annual[da_annual['year']==2022]['avg_price'].values[0], 1)}</div>
                <div class="label">2022 Avg DA Price<br>(Energy Crisis)</div>
            </div>
            <div class="stat">
                <div class="value">{neg_hours_real.get(2025, 'N/A')}</div>
                <div class="label">Negative Price<br>Hours (2025)</div>
            </div>
            <div class="stat green">
                <div class="value">{round(spread_by_year[spread_by_year['year']==2025]['avg_spread'].values[0], 0) if 2025 in spread_by_year['year'].values else 'N/A'}</div>
                <div class="label">2025 Avg Daily<br>Spread EUR/MWh</div>
            </div>
        </div>
    </div>
    <div class="card">
        <div class="stat-grid">
            <div class="stat">
                <div class="value">{round(load_annual[load_annual['year']==2025]['avg_load'].values[0]/1000, 2) if 2025 in load_annual['year'].values else 'N/A'}</div>
                <div class="label">2025 Avg Load<br>GW</div>
            </div>
            <div class="stat red">
                <div class="value">{round(load_annual[load_annual['year']==2025]['max_load'].values[0]/1000, 2) if 2025 in load_annual['year'].values else 'N/A'}</div>
                <div class="label">2025 Peak<br>Load GW</div>
            </div>
            <div class="stat green">
                <div class="value">535</div>
                <div class="label">BESS Installed<br>2025 (MW)</div>
            </div>
            <div class="stat solar">
                <div class="value">1,700</div>
                <div class="label">BESS Pipeline<br>Target (MW)</div>
            </div>
        </div>
    </div>
</div>

<!-- ========== REVENUE PROJECTION HEADLINE ========== -->
<div class="section-divider"><h2>1. BESS Revenue Projection (2h System)</h2></div>

<div class="grid">
    <div class="card card-full">
        <div class="highlight-box">
            <h3 style="color:#856404; margin-top:0;">Window of Opportunity: 2025—2027</h3>
            <p>Projects operational by 2025-2026 capture extraordinary revenues of ~€390-570/kW/yr driven by post-BRELL scarcity pricing. After 2027, expect 55-70% revenue compression as 1,200+ MW BESS pipeline enters the market.</p>
        </div>
    </div>
</div>

<div class="grid">
    <div class="card card-full">
        <h2>Multi-Market Revenue by Source (EUR/kW/yr, 2h BESS)</h2>
        <div id="chart_revenue_stacked" style="width:100%; min-height:450px;"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Revenue Projection (EUR/kW/yr, 2h system)</h2>
        <p style="font-size:0.85em; color:#666; margin-top:0">Multi-market combined (aFRR 40%, FCR 20%, DA 25%, mFRR 5%, Imbalance 10%).</p>
        <table>
            <tr><th>Year</th><th>2024</th><th>2025</th><th>2026E</th><th>2027E</th><th>2028E</th><th>2029E</th><th>2030E</th></tr>
            <tr><td>EUR/MW/yr</td><td>74,680</td><td>571,719</td><td>387,309</td><td>272,453</td><td>220,099</td><td>193,922</td><td>182,182</td></tr>
            <tr style="font-weight:bold; background:#D4EDDA"><td>EUR/kW/yr</td><td>75</td><td>572</td><td>387</td><td>272</td><td>220</td><td>194</td><td>182</td></tr>
            <tr><td>Compression</td><td>100%</td><td>100%</td><td>65%</td><td>45%</td><td>35%</td><td>30%</td><td>28%</td></tr>
            <tr><td>BESS MW</td><td>250</td><td>454</td><td>700</td><td>1,200</td><td>1,500</td><td>1,800</td><td>2,000</td></tr>
        </table>
    </div>
    <div class="card">
        <h2>Revenue Composition Shift</h2>
        <table>
            <tr><th>Stream</th><th>2024</th><th>2025</th><th>2026E</th><th>2027E</th><th>2028E</th><th>2030E</th></tr>
            <tr><td>DA Arbitrage</td><td>31%</td><td>5%</td><td>6%</td><td>7%</td><td>7%</td><td>8%</td></tr>
            <tr><td>aFRR</td><td>8%</td><td>58%</td><td>56%</td><td>54%</td><td>54%</td><td>54%</td></tr>
            <tr><td>FCR</td><td>0%</td><td>9%</td><td>10%</td><td>11%</td><td>12%</td><td>10%</td></tr>
            <tr><td>mFRR</td><td>41%</td><td>25%</td><td>25%</td><td>24%</td><td>23%</td><td>24%</td></tr>
            <tr><td>Imbalance</td><td>20%</td><td>3%</td><td>3%</td><td>4%</td><td>4%</td><td>5%</td></tr>
        </table>
        <p style="font-size:0.85em; color:#666; margin-top:10px">aFRR dominates post-BRELL. Pre-2025, mFRR and DA were primary revenue drivers.</p>
    </div>
</div>

<!-- ========== DA PRICES ========== -->
<div class="section-divider"><h2>2. Day-Ahead Electricity Prices</h2></div>

<div class="grid">
    <div class="card">
        <h2>Annual DA Price Summary</h2>
        <table>
            <tr><th>Year</th><th>Avg</th><th>Min</th><th>Max</th><th>Std Dev</th><th>Avg Spread</th><th>Neg Hrs</th></tr>
"""

for _, row in da_annual.iterrows():
    yr = int(row['year'])
    nh = neg_hours_real.get(yr, 0)
    spr = spread_by_year[spread_by_year['year']==yr]['avg_spread'].values
    spread_val = f"{spr[0]:.0f}" if len(spr) > 0 else "—"
    html += f"""            <tr><td><strong>{yr}</strong></td>
                <td>{row['avg_price']:.1f}</td><td>{row['min_price']:.0f}</td>
                <td>{row['max_price']:.0f}</td><td>{row['std_price']:.1f}</td>
                <td>{spread_val}</td><td>{nh}</td></tr>\n"""

html += """        </table>
    </div>
    <div class="card">
        <h2>Daily Price Spread — Arbitrage Potential</h2>
        <div id="chart_spread" class="chart" style="min-height:350px"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Hourly Price Profile by Year</h2>
        <div id="chart_hourly" class="chart" style="min-height:380px"></div>
    </div>
    <div class="card">
        <h2>Price Distribution — Neg/Pos Hours</h2>
        <div id="chart_neg_hours" class="chart" style="min-height:380px"></div>
    </div>
</div>

<!-- ========== IMBALANCE ========== -->
<div class="section-divider"><h2>2. Imbalance Prices</h2></div>

<div class="grid">
    <div class="card card-full">
        <h2>Imbalance Settlement Prices (EUR/MWh) — Long vs Short</h2>
        <div id="chart_imbalance" class="chart"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Annual Imbalance Summary (2021 — Sep 2024)</h2>
        <table>
            <tr><th>Year</th><th>Avg Long</th><th>Avg Short</th><th>Min Long</th><th>Max Short</th><th>Avg Spread</th></tr>
"""

for year, row in imb_annual.iterrows():
    html += f"""            <tr><td><strong>{int(year)}</strong></td>
                <td>{row['avg_long']:.1f}</td><td>{row['avg_short']:.1f}</td>
                <td>{row['min_long']:.1f}</td><td>{row['max_short']:.1f}</td>
                <td>{row['avg_spread']:.1f}</td></tr>\n"""

html += """        </table>
        <div class="insight">
            <strong>Key:</strong> Imbalance settlement shifted from 60-min to 15-min periods in 2024 (EU regulation).
            Lithuania joined PICASSO (pan-European aFRR) in March 2025.
        </div>
    </div>
    <div class="card">
        <h2>aFRR Market Indicators</h2>
        <table>
            <tr><th>Parameter</th><th>2024</th><th>2025</th><th>2026E</th></tr>
            <tr><td>aFRR Up Req (MW)</td><td>60</td><td>80</td><td>96-120</td></tr>
            <tr><td>aFRR Down Req (MW)</td><td>60</td><td>80</td><td>104-120</td></tr>
            <tr><td>Est. Capacity Price Up (EUR/MW/h)</td><td>20</td><td>25</td><td>30</td></tr>
            <tr><td>BESS Share of aFRR</td><td>~25%</td><td>~40%</td><td>~55%</td></tr>
            <tr><td>Ancillary Cost (ct/kWh)</td><td>1.31</td><td>~1.5</td><td>~1.8</td></tr>
        </table>
        <div class="insight">
            <strong>Insight:</strong> Ancillary service costs surged 5.5x in 2024 vs 2023. BESS is increasingly the lowest-cost provider of aFRR.
        </div>
    </div>
</div>

<!-- ========== LOAD & GENERATION ========== -->
<div class="section-divider"><h2>3. Electricity Load & Generation</h2></div>

<div class="grid">
    <div class="card card-full">
        <h2>Actual Electricity Load (MW) — ENTSO-E Data</h2>
        <div id="chart_load_annual" class="chart"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Annual Load Summary</h2>
        <table>
            <tr><th>Year</th><th>Avg (MW)</th><th>Min (MW)</th><th>Peak (MW)</th><th>Est TWh</th></tr>
"""

for _, row in load_annual.iterrows():
    html += f"""            <tr><td><strong>{int(row['year'])}</strong></td>
                <td>{row['avg_load']:.0f}</td><td>{row['min_load']:.0f}</td>
                <td>{row['max_load']:.0f}</td><td>{row['twh']:.1f}</td></tr>\n"""

html += """        </table>
    </div>
    <div class="card">
        <h2>Generation by Source (TWh/yr) — ENTSO-E</h2>
        <div id="chart_gen_stack" class="chart" style="min-height:400px"></div>
    </div>
</div>

<div class="grid">
    <div class="card card-full">
        <h2>Cross-border Electricity Flows (TWh)</h2>
        <div id="chart_flows" class="chart"></div>
    </div>
</div>

<!-- ========== INSTALLED CAPACITY ========== -->
<div class="section-divider"><h2>4. Installed Capacity — Historical & Forecast</h2></div>

<div class="grid">
    <div class="card card-full">
        <h2>Cumulative Installed Capacity (MW)</h2>
        <div id="chart_capacity" class="chart"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Capacity Data (MW)</h2>
        <table>
            <tr><th>Year</th><th>Wind</th><th>Solar PV</th><th>BESS (MW)</th><th>BESS (MWh)</th><th>Fossil</th><th>Total RES</th></tr>
"""

for i, yr in enumerate(years_cap):
    fc = ' class="forecast"' if yr >= 2026 else ''
    html += f"""            <tr{fc}><td><strong>{yr}</strong></td>
                <td>{wind_cumul[i]:,}</td><td>{pv_cumul[i]:,}</td>
                <td>{bess_power[i]:,}</td><td>{bess_energy[i]:,}</td>
                <td>{fossil_mw[i]:,}</td><td>{wind_cumul[i]+pv_cumul[i]:,}</td></tr>\n"""

html += f"""        </table>
        <p style="font-size:0.85em; color:#856404; margin-top:8px;">
            Yellow rows = Forecast. Sources: LVEA, Litgrid, PV Magazine, IRENA, NECP
        </p>
    </div>
    <div class="card">
        <h2>Key Capacity Milestones</h2>
        <div class="bullet">Wind + Solar reached <strong>3 GW</strong> by end 2025 (Litgrid confirmed)</div>
        <div class="bullet">Solar grew from 250 MW (2021) to 2,800 MW (2025) — <strong>11x in 4 years</strong></div>
        <div class="bullet">BESS: from 0 → 535 MW operational by mid-2025</div>
        <div class="bullet">Government BESS target: <strong>1.7 GW / 4 GWh</strong> (€197mn+ allocated)</div>
        <div class="bullet">NECP 2030: 5.1 GW solar, ~4 GW wind (incl. 0.7 GW offshore)</div>
        <div class="bullet">Fossil: Elektrėnai complex 1,800 MW (455 MW CCGT active, rest reserve)</div>
        <div class="bullet">Ignitis: 291 MW / 582 MWh BESS coming online 2027 (€130mn)</div>
        <div class="bullet">Trina: 90 MW / 180 MWh across 3 projects (mid-2026)</div>
    </div>
</div>

<!-- ========== BESS SATURATION ========== -->
<div class="section-divider"><h2>5. BESS Market Saturation Analysis</h2></div>

<div class="grid">
    <div class="card">
        <h2>Revenue vs Saturation</h2>
        <div id="chart_saturation" class="chart" style="min-height:380px"></div>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>Saturation Metrics</h2>
        <table>
            <tr><th>Metric</th><th>2024</th><th>2025</th><th>2026E</th><th>2027E</th><th>2028E</th><th>2030E</th></tr>
            <tr><td>BESS / Peak Demand</td><td>12%</td><td>25%</td><td>36%</td><td>52%</td><td>63%</td><td>67%</td></tr>
            <tr><td>BESS / Avg Load</td><td>17%</td><td>35%</td><td>49%</td><td>70%</td><td>82%</td><td>88%</td></tr>
            <tr><td>BESS exceeds aFRR mkt</td><td>No</td><td>No</td><td>No</td>
                <td><span class="tag tag-orange">Yes</span></td>
                <td><span class="tag tag-red">Yes</span></td>
                <td><span class="tag tag-red">Yes</span></td></tr>
            <tr><td>Revenue Cannibalization</td><td>0%</td><td>5%</td><td>10%</td><td>18%</td><td>25%</td><td>35%</td></tr>
            <tr><td>Signal</td>
                <td><span class="tag tag-green">Low</span></td>
                <td><span class="tag tag-yellow">Low-Med</span></td>
                <td><span class="tag tag-yellow">Medium</span></td>
                <td><span class="tag tag-orange">Med-High</span></td>
                <td><span class="tag tag-red">High</span></td>
                <td><span class="tag tag-red">Saturated</span></td></tr>
        </table>
    </div>
    <div class="card">
        <h3>Bull Case (Favorable)</h3>
        <div class="bullet">Russia disconnection → structural price volatility increase</div>
        <div class="bullet">Load growth 4%/yr (data centers, EVs, heat pumps)</div>
        <div class="bullet">0.7 GW offshore wind by 2030 → large balancing need</div>
        <div class="bullet">PICASSO cross-border aFRR revenue opportunity</div>
        <div class="bullet">€197mn+ government support (largest EU BESS procurement)</div>

        <h3 style="margin-top:20px">Bear Case (Risks)</h3>
        <div class="bullet">1.7 GW/4 GWh pipeline → BESS/Peak at 67% by 2029</div>
        <div class="bullet">aFRR market small (~120 MW) vs BESS pipeline (1,700 MW)</div>
        <div class="bullet">Revenue cannibalization 25-35% by 2028-2030</div>
        <div class="bullet">Small market (13-18 TWh) limits absolute opportunity</div>
    </div>
</div>

<!-- ========== MARKET OVERVIEW ========== -->
<div class="section-divider"><h2>6. Market Overview & Key Developments</h2></div>

<div class="grid">
    <div class="card">
        <h2>Grid Disconnection from Russia</h2>
        <div class="bullet">8 Feb 2025: Baltic states disconnected from BRELL grid</div>
        <div class="bullet">9 Feb 2025: Synchronized with Continental Europe via LitPol Link</div>
        <div class="bullet">Price spike to €325/MWh peak, €128/MWh daily avg on day of disconnect</div>
        <div class="bullet">Ended 50+ years of Soviet/Russian grid dependency</div>
        <div class="bullet">Structural increase in flexibility need → bullish BESS</div>
    </div>
    <div class="card">
        <h2>Interconnections</h2>
        <table>
            <tr><th>Link</th><th>Capacity</th><th>Status</th></tr>
            <tr><td>NordBalt (→Sweden)</td><td>700 MW HVDC</td><td>Operational (2015)</td></tr>
            <tr><td>LitPol Link (→Poland)</td><td>500 MW</td><td>AC mode (Feb 2025)</td></tr>
            <tr><td>Latvia Interconnection</td><td>~1,200 MW</td><td>Operational</td></tr>
            <tr><td>Harmony Link (→Poland)</td><td>700 MW HVDC</td><td>Under development</td></tr>
            <tr><td>Offshore Hub (→DE/LV)</td><td>2,000 MW</td><td>Planned (post-2035)</td></tr>
        </table>
    </div>
</div>

<div class="grid">
    <div class="card">
        <h2>BESS Project Pipeline</h2>
        <table>
            <tr><th>Developer</th><th>Power</th><th>Energy</th><th>COD</th></tr>
            <tr><td>Litgrid (Fluence)</td><td>200 MW</td><td>200 MWh</td><td>2022 ✅</td></tr>
            <tr><td>Energy Cells (aFRR)</td><td>40 MW</td><td>N/A</td><td>2023 ✅</td></tr>
            <tr><td>Trina Storage (3 sites)</td><td>90 MW</td><td>180 MWh</td><td>Mid-2026</td></tr>
            <tr><td>European Energy (Anykščiai)</td><td>N/A</td><td>65 MWh</td><td>Feb 2026</td></tr>
            <tr><td>Ignitis Group (3 sites)</td><td>291 MW</td><td>582 MWh</td><td>2027</td></tr>
            <tr><td>Aura/Balancy (Kaišiadorys)</td><td>50 MW</td><td>100 MWh</td><td>TBD</td></tr>
            <tr style="font-weight:bold; background:#D4EDDA"><td>Govt Target</td><td>1,700 MW</td><td>4,000 MWh</td><td>~2029</td></tr>
        </table>
    </div>
    <div class="card">
        <h2>Regulatory & Policy</h2>
        <div class="bullet">€197mn+ BESS support allocated (2x original budget)</div>
        <div class="bullet">Additional €45mn round announced Oct 2025</div>
        <div class="bullet">100% renewable electricity target by 2030</div>
        <div class="bullet">Net electricity exporter target by 2028</div>
        <div class="bullet">EU ETS2 introduction 2027 → upward price pressure</div>
        <div class="bullet">15-min imbalance settlement (from 60-min)</div>
        <div class="bullet">PICASSO integration for cross-border aFRR</div>
        <div class="bullet">Energy independence = national security priority post-Ukraine</div>
    </div>
</div>

</div><!-- container -->

<footer>
    <p>Lithuania BESS Market Analysis — Birdview Energy — {datetime.now().strftime('%B %Y')}</p>
    <p>Data sources: ENTSO-E Transparency Platform, Nord Pool, Litgrid, IRENA, LVEA, IEA, PV Magazine, CEE Energy News, ESS News</p>
    <p style="margin-top:8px; font-size:0.85em;">Forecasts are estimates based on published targets, market trends, and industry analysis. Actual outcomes may vary.</p>
</footer>

<script>
// ==================== CHART DATA ====================
const monthlyPrices = {json.dumps(monthly_prices)};
const hourlyData = {json.dumps(hourly_data)};
const genData = {json.dumps(gen_chart_data)};
const spreadData = {json.dumps({int(k): v for k, v in spread_chart.items()})};
const negHoursData = {json.dumps({int(k): int(v) for k, v in neg_hours_real.items()})};

// Load data
const loadData = {json.dumps({int(k): v for k, v in load_chart.items()})};
const flowData = {json.dumps({int(k): {kk: round(float(vv), 2) for kk, vv in v.items()} for k, v in flow_chart.items()})};

// Capacity
const yearsCap = {json.dumps(years_cap)};
const windCumul = {json.dumps(wind_cumul)};
const pvCumul = {json.dumps(pv_cumul)};
const bessPower = {json.dumps(bess_power)};
const bessEnergy = {json.dumps(bess_energy)};
const fossilMW = {json.dumps(fossil_mw)};
const bessRevenue = {json.dumps(bess_revenue)};
const bessSatPct = {json.dumps(bess_saturation_pct)};

const plotlyConfig = {{responsive: true, displayModeBar: false}};
const colors = {{
    primary: '#1F4E79', accent: '#2E75B6', green: '#70AD47',
    solar: '#FFC000', red: '#C00000', purple: '#7B2D8E',
    orange: '#ED7D31', gray: '#A5A5A5'
}};

// ==================== 0. REVENUE STACKED BAR ====================
(() => {{
    const years = ['2024', '2025', '2026E', '2027E', '2028E', '2029E', '2030E'];
    const da =        [23, 27, 23, 19, 16, 16, 14];
    const afrr =      [6, 334, 218, 148, 118, 103, 98];
    const fcr =       [0, 52, 39, 31, 26, 22, 18];
    const mfrr =      [31, 144, 95, 64, 51, 44, 43];
    const imbalance = [15, 15, 12, 10, 9, 9, 9];

    Plotly.newPlot('chart_revenue_stacked', [
        {{x: years, y: da, name: 'DA Arbitrage', type: 'bar',
         marker: {{color: '#2E75B6'}}, hovertemplate: 'DA: %{{y}} EUR/kW/yr<extra></extra>'}},
        {{x: years, y: afrr, name: 'aFRR', type: 'bar',
         marker: {{color: '#70AD47'}}, hovertemplate: 'aFRR: %{{y}} EUR/kW/yr<extra></extra>'}},
        {{x: years, y: fcr, name: 'FCR', type: 'bar',
         marker: {{color: '#FFC000'}}, hovertemplate: 'FCR: %{{y}} EUR/kW/yr<extra></extra>'}},
        {{x: years, y: mfrr, name: 'mFRR', type: 'bar',
         marker: {{color: '#ED7D31'}}, hovertemplate: 'mFRR: %{{y}} EUR/kW/yr<extra></extra>'}},
        {{x: years, y: imbalance, name: 'Imbalance', type: 'bar',
         marker: {{color: '#7B2D8E'}}, hovertemplate: 'Imbalance: %{{y}} EUR/kW/yr<extra></extra>'}}
    ], {{
        title: {{text: 'Revenue by Market Source — 2h BESS (EUR/kW/yr)', font: {{size: 16}}}},
        xaxis: {{type: 'category'}},
        yaxis: {{title: 'EUR/kW/yr', gridcolor: '#E0E0E0'}},
        barmode: 'stack',
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.18, x: 0.5, xanchor: 'center'}},
        margin: {{t: 50, b: 70, l: 60, r: 20}},
        annotations: [
            {{x: '2025', y: 572, text: '<b>€572/kW/yr</b>', showarrow: true,
             arrowhead: 0, ay: -30, font: {{size: 13, color: '#1F4E79'}}}},
            {{x: '2030E', y: 182, text: '<b>€182/kW/yr</b>', showarrow: true,
             arrowhead: 0, ay: -30, font: {{size: 12, color: '#C00000'}}}},
            {{x: '2024', y: 75, text: 'Pre-BRELL', showarrow: false,
             yshift: 15, font: {{size: 10, color: '#666'}}}},
        ],
        shapes: [{{type: 'line', x0: 0.5, x1: 0.5, y0: 0, y1: 620,
                  line: {{color: '#C00000', width: 2, dash: 'dash'}}}}],
    }}, plotlyConfig);
}})();

// ==================== 1. SPREAD CHART ====================
(() => {{
    const years = Object.keys(spreadData).sort().filter(y => Number(y) >= 2023);
    Plotly.newPlot('chart_spread', [
        {{x: years, y: years.map(y => spreadData[y]?.avg_spread?.toFixed(1)), name: 'Avg Spread', type: 'bar', marker: {{color: colors.accent}}}},
        {{x: years, y: years.map(y => spreadData[y]?.median_spread?.toFixed(1)), name: 'Median', type: 'bar', marker: {{color: colors.green}}}},
        {{x: years, y: years.map(y => spreadData[y]?.p90_spread?.toFixed(1)), name: 'P90 Spread', type: 'scatter', mode: 'lines+markers', marker: {{color: colors.red}}, yaxis: 'y'}}
    ], {{
        title: 'Daily DA Price Spread (Max-Min)',
        yaxis: {{title: 'EUR/MWh', gridcolor: '#E0E0E0'}},
        barmode: 'group', plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.2}}, margin: {{t: 40, b: 60}}
    }}, plotlyConfig);
}})();

// ==================== 3. HOURLY PROFILE ====================
(() => {{
    const years = Object.keys(hourlyData).sort();
    const selYears = years.filter(y => [2021,2022,2024,2025].includes(Number(y)));
    const traces = selYears.map(year => {{
        const hours = Array.from({{length: 24}}, (_, i) => i);
        return {{
            x: hours, y: hours.map(h => hourlyData[year][h] || null),
            name: year, type: 'scatter', mode: 'lines', line: {{width: 2.5}}
        }};
    }});
    Plotly.newPlot('chart_hourly', traces, {{
        title: 'Hourly Price Profile (Average EUR/MWh)',
        xaxis: {{title: 'Hour of Day', dtick: 2, gridcolor: '#E0E0E0'}},
        yaxis: {{title: 'EUR/MWh', gridcolor: '#E0E0E0'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.2}}, margin: {{t: 40, b: 60}}
    }}, plotlyConfig);
}})();

// ==================== 4. NEG HOURS ====================
(() => {{
    const years = Object.keys(negHoursData).sort();
    Plotly.newPlot('chart_neg_hours', [{{
        x: years, y: years.map(y => negHoursData[y]),
        type: 'bar', marker: {{color: years.map(y => negHoursData[y] > 100 ? colors.red : colors.accent)}}
    }}], {{
        title: 'Negative Price Hours per Year',
        yaxis: {{title: 'Hours', gridcolor: '#E0E0E0'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        showlegend: false, margin: {{t: 40, b: 40}}
    }}, plotlyConfig);
}})();

// ==================== 5. IMBALANCE MONTHLY ====================
(() => {{
    // We'll create a simple annual bar chart from imb data
    const imbYears = {json.dumps([int(y) for y in imb_annual.index])};
    const avgLong = {json.dumps([round(float(v), 1) for v in imb_annual['avg_long']])};
    const avgShort = {json.dumps([round(float(v), 1) for v in imb_annual['avg_short']])};
    const maxShort = {json.dumps([round(float(v), 1) for v in imb_annual['max_short']])};
    const minLong = {json.dumps([round(float(v), 1) for v in imb_annual['min_long']])};

    Plotly.newPlot('chart_imbalance', [
        {{x: imbYears, y: avgLong, name: 'Avg Long (surplus)', type: 'bar', marker: {{color: colors.green}}}},
        {{x: imbYears, y: avgShort, name: 'Avg Short (deficit)', type: 'bar', marker: {{color: colors.red}}}},
        {{x: imbYears, y: maxShort, name: 'Max Short', type: 'scatter', mode: 'markers', marker: {{color: colors.red, size: 10, symbol: 'triangle-up'}}}},
        {{x: imbYears, y: minLong, name: 'Min Long', type: 'scatter', mode: 'markers', marker: {{color: colors.green, size: 10, symbol: 'triangle-down'}}}}
    ], {{
        title: 'Annual Imbalance Prices (EUR/MWh)',
        yaxis: {{title: 'EUR/MWh', gridcolor: '#E0E0E0'}},
        barmode: 'group', plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.2}}, margin: {{t: 40, b: 60}}
    }}, plotlyConfig);
}})();

// ==================== 6. LOAD ANNUAL ====================
(() => {{
    const years = Object.keys(loadData).sort();
    Plotly.newPlot('chart_load_annual', [
        {{x: years, y: years.map(y => (loadData[y]?.avg_load / 1000).toFixed(2)), name: 'Avg Load (GW)', type: 'bar', marker: {{color: colors.accent}}}},
        {{x: years, y: years.map(y => (loadData[y]?.max_load / 1000).toFixed(2)), name: 'Peak Load (GW)', type: 'scatter', mode: 'lines+markers', marker: {{color: colors.red, size: 8}}, line: {{width: 2.5}}}},
        {{x: years, y: years.map(y => (loadData[y]?.min_load / 1000).toFixed(2)), name: 'Min Load (GW)', type: 'scatter', mode: 'lines+markers', marker: {{color: colors.green, size: 8}}, line: {{width: 2.5, dash: 'dot'}}}}
    ], {{
        title: 'Annual Electricity Load (GW)',
        yaxis: {{title: 'GW', gridcolor: '#E0E0E0'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.15}}, margin: {{t: 40, b: 50}}
    }}, plotlyConfig);
}})();

// ==================== 7. GENERATION BY TYPE ====================
(() => {{
    const mainTypes = ['Wind', 'Solar', 'Gas', 'Hydro', 'Biomass', 'Pumped Storage (Gen)', 'Battery Storage'];
    const typeColors = {{Wind: colors.accent, Solar: colors.solar, Gas: colors.gray,
                        Hydro: '#00B0F0', Biomass: colors.green, 'Pumped Storage (Gen)': colors.purple,
                        'Battery Storage': colors.orange}};
    const years = Object.keys(genData['Wind'] || {{}}).sort();
    const traces = mainTypes.filter(t => genData[t]).map(t => ({{
        x: years, y: years.map(y => genData[t][y]?.toFixed(3) || 0),
        name: t, type: 'bar', marker: {{color: typeColors[t] || '#999'}}
    }}));
    Plotly.newPlot('chart_gen_stack', traces, {{
        title: 'Generation by Source (TWh)',
        yaxis: {{title: 'TWh', gridcolor: '#E0E0E0'}},
        barmode: 'stack', plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.2}}, margin: {{t: 40, b: 60}}
    }}, plotlyConfig);
}})();

// ==================== 8. CROSS-BORDER FLOWS ====================
(() => {{
    const years = Object.keys(flowData).sort();
    Plotly.newPlot('chart_flows', [
        {{x: years, y: years.map(y => flowData[y]?.import_twh), name: 'Import (TWh)', type: 'bar', marker: {{color: colors.accent}}}},
        {{x: years, y: years.map(y => -flowData[y]?.export_twh), name: 'Export (TWh)', type: 'bar', marker: {{color: colors.green}}}},
        {{x: years, y: years.map(y => flowData[y]?.net_import_twh), name: 'Net Import', type: 'scatter', mode: 'lines+markers', line: {{color: colors.red, width: 3}}, marker: {{size: 8}}}}
    ], {{
        title: 'Cross-border Electricity Flows (TWh)',
        yaxis: {{title: 'TWh', gridcolor: '#E0E0E0'}},
        barmode: 'relative', plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.15}}, margin: {{t: 40, b: 50}}
    }}, plotlyConfig);
}})();

// ==================== 9. INSTALLED CAPACITY ====================
(() => {{
    const divider = yearsCap.indexOf(2026);
    Plotly.newPlot('chart_capacity', [
        {{x: yearsCap, y: windCumul, name: 'Wind (MW)', type: 'bar', marker: {{color: colors.accent}}}},
        {{x: yearsCap, y: pvCumul, name: 'Solar PV (MW)', type: 'bar', marker: {{color: colors.solar}}}},
        {{x: yearsCap, y: bessPower, name: 'BESS (MW)', type: 'bar', marker: {{color: colors.green}}}},
        {{x: yearsCap, y: fossilMW, name: 'Fossil (MW)', type: 'scatter', mode: 'lines+markers',
          line: {{color: colors.gray, width: 2, dash: 'dot'}}, marker: {{size: 6}}}}
    ], {{
        title: 'Installed Capacity (MW) — Historical + Forecast',
        yaxis: {{title: 'MW', gridcolor: '#E0E0E0'}},
        barmode: 'stack', plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.15}}, margin: {{t: 40, b: 50}},
        shapes: [{{type: 'line', x0: 2025.5, x1: 2025.5, y0: 0, y1: 12000,
                   line: {{color: '#C00000', width: 2, dash: 'dash'}}}}],
        annotations: [{{x: 2025.5, y: 11500, text: 'Forecast →', showarrow: false,
                        font: {{color: '#C00000', size: 12}}}}]
    }}, plotlyConfig);
}})();

// ==================== 10. SATURATION ====================
(() => {{
    Plotly.newPlot('chart_saturation', [
        {{x: yearsCap, y: bessRevenue, name: 'Revenue (EUR/kW/yr)', type: 'bar',
          marker: {{color: yearsCap.map((y,i) => bessRevenue[i] > 300 ? colors.green : bessRevenue[i] > 200 ? colors.solar : colors.red)}}}},
        {{x: yearsCap, y: bessSatPct, name: 'BESS/Peak Demand (%)', type: 'scatter', mode: 'lines+markers',
          line: {{color: colors.red, width: 3}}, marker: {{size: 8}}, yaxis: 'y2'}}
    ], {{
        title: 'Revenue vs Saturation',
        yaxis: {{title: 'EUR/kW/yr', gridcolor: '#E0E0E0', side: 'left'}},
        yaxis2: {{title: 'BESS/Peak %', overlaying: 'y', side: 'right', gridcolor: 'transparent'}},
        plot_bgcolor: 'white', paper_bgcolor: 'white',
        legend: {{orientation: 'h', y: -0.15}}, margin: {{t: 40, b: 50}},
        shapes: [{{type: 'rect', x0: 2024.5, x1: 2027.5, y0: 0, y1: 620,
                   fillcolor: 'rgba(112,173,71,0.1)', line: {{width: 0}}}}],
        annotations: [{{x: 2026, y: 590, text: 'Golden Window', showarrow: false,
                        font: {{color: colors.green, size: 13, bold: true}}}}]
    }}, plotlyConfig);
}})();

</script>
</body>
</html>"""

# Reorder: move DA Prices (section 1) and Imbalance (section 2) to bottom of page
da_marker = '<!-- ========== DA PRICES ========== -->'
load_marker = '<!-- ========== LOAD & GENERATION ========== -->'
footer_marker = '</div><!-- container -->'

if da_marker in html and load_marker in html:
    da_start = html.find(da_marker)
    load_start = html.find(load_marker)
    sections_1_2 = html[da_start:load_start]
    html = html[:da_start] + html[load_start:]
    html = html.replace(footer_marker, sections_1_2 + '\n' + footer_marker)

# Write HTML file
html_path = f"{OUT_DIR}/Lithuania_BESS_Market_Report.html"
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\nHTML report saved to: {html_path}")
print("Done! Open in browser to view interactive charts.")
